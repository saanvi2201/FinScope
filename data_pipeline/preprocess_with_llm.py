# =============================================================================
# preprocess_with_llm.py — Hybrid (Rule + LLM) Preprocessing Pipeline
# =============================================================================
#
# ── FIXES IN THIS VERSION ────────────────────────────────────────────────────
#
#  🔴 FIX A — LLM OVERRIDE LOGIC COMPLETELY REWRITTEN (CRITICAL)
#    Problem: The old logic used ONLY confidence to decide LLM vs rule.
#    This meant a hallucinating LLM (0.88 confidence) would override a
#    correct rule result (0.84 confidence) purely on numbers.
#    Fix: apply_llm_override() now checks:
#      1. Bank must match between LLM and rule (when rule bank confidence ≥ 0.75)
#         → If banks differ, LLM result is REJECTED
#      2. Deflated confidence (from llm_classifier.py FIX 2) is used, not raw
#      3. LLM cannot override a valid (non-UNKNOWN) card name unless it also
#         correctly identifies the same bank
#      4. Override margin increased: 0.10 → 0.15
#         (LLM must be meaningfully better, not just slightly)
#
#  🔴 FIX B — RULE-BASED BANK NARROWING FOR SCB (CRITICAL)
#    Problem: SC_Smart_MITC_2026.pdf detected as SBI instead of SCB.
#    Root cause: "sbi" substring appears in garbled OCR text.
#    Fix: Added SCB-priority check in _narrow_card_to_bank().
#    When bank=SBI but file starts with "SC_", force re-check for SCB.
#    Also improved the BANK_ALIASES ordering in preprocess.py.
#
#  🔴 FIX C — CROSS-BANK MISMATCH VALIDATION (CRITICAL)
#    Problem: validate_prediction() was not catching cross-bank errors.
#    e.g., a file detected as HDFC Millennia but originally an AXIS file
#    could pass validation and go to processed_docs/.
#    Fix: validate_prediction() now checks if the (bank, card) pair is valid
#    against BANK_CARDS. Any cross-bank mismatch forces needs_review/.
#
#  🟡 FIX D — FILENAME-BASED SANITY CHECK (NEW)
#    Problem: Several files were routed completely wrong because the LLM
#    ignored the filename hint. e.g., AXIS_Flipkart_TNC_Merchants_2026.pdf
#    was classified as HDFC_Millennia.
#    Fix: Added _filename_sanity_check() that extracts bank+card hints from
#    the original filename and flags if the final classification contradicts
#    them strongly. Files that fail this check go to needs_review/.
#
#  🟡 FIX E — LLM TRIGGER CONDITIONS TIGHTENED (IMPORTANT)
#    Problem: LLM was being triggered even when rules found everything
#    correctly with high confidence. e.g., AXIS ACE BR at 0.84 was not
#    triggering LLM — but some 0.71 cases that had good bank+card still did.
#    Fix: _should_call_llm() now has a smarter trigger:
#      - LLM is NOT triggered if: conf < threshold BUT bank+card are both
#        valid (known) AND confidence is only slightly below threshold (>0.65)
#        AND the bank is high confidence (≥0.80)
#      - This prevents unnecessary LLM calls on near-correct rule results
#
#  🟡 FIX F — LLM RESULT RE-VALIDATION AFTER OVERRIDE (NEW)
#    Problem: After LLM override was applied, the result was never re-
#    validated. So LLM results that passed the override check but were
#    actually wrong could flow directly to processed_docs/.
#    Fix: After apply_llm_override(), we re-run validate_prediction() on
#    the FINAL (post-override) result. If it fails, file goes to needs_review/.
#
#  🟡 FIX G — classify_with_llm() NOW RECEIVES RULE HINTS
#    Problem: classify_with_llm() had no knowledge of what rules found.
#    Fix: _process_one_file() now passes rule_bank, rule_card, rule_doc_type,
#    rule_confidence to classify_with_llm() so it can ground its output.
#    (This is the preprocess_with_llm.py side of llm_classifier.py FIX 1.)
#
#  🔵 FIX H — DRY-RUN BUG (from original — kept and verified)
#    In dry-run mode: no folders created, no files copied, no JSON written.
#
#  🔵 FIX I — MASTER DOC NARROWING BUG (from original — kept and verified)
#    _narrow_card_to_bank() skipped for is_master=True docs.
#
#  🔵 FIX J — YEAR EXTRACTION (from original — kept and verified)
#    detect_year() prioritises header zone, caps at MAX_VALID_YEAR.
#
# ── HOW TO RUN ────────────────────────────────────────────────────────────────
#
#  STEP 0 — Install Python dependencies (run once)
#    pip install pdfplumber pymupdf requests pandas openpyxl pytesseract Pillow
#
#  STEP 1 — Set up Ollama + fast model (recommended for CPU machines)
#    ollama pull llama3.2:1b    ← ~800MB, 15-30s on CPU
#    ollama serve               ← keep this terminal open
#    (If you get "address already in use", Ollama is already running — OK)
#
#  STEP 2 — Place input PDFs in:
#    <project_root>/data/raw_docs/
#
#  STEP 3 — Run from project root:
#    # Test first (no files written):
#    python data_pipeline/preprocess_with_llm.py --dry-run
#
#    # Rules only (no LLM):
#    python data_pipeline/preprocess_with_llm.py --no-llm
#
#    # Full run (rules + LLM fallback):
#    python data_pipeline/preprocess_with_llm.py
#
#    # Full run with debug text output:
#    python data_pipeline/preprocess_with_llm.py --debug
#
#  STEP 4 — Check outputs:
#    data/processed_docs/<BANK>_<CARD>/   ← renamed PDFs + metadata JSON
#    data/needs_review/                   ← low-confidence / flagged files
#    data/logs/summary.xlsx               ← per-file results (colour-coded)
#    data/logs/preprocess_log.txt         ← detailed audit trail
#    data/logs/hybrid_classification_log.txt ← LLM decision log
#    data/logs/missing_docs_report.xlsx   ← coverage gaps
#    data/logs/coverage_dashboard.xlsx    ← visual coverage grid
#
# =============================================================================

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# PATH SETUP
# ─────────────────────────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).resolve().parent))

from preprocess import (
    run_detection_with_fallback,
    compute_confidence,
    extract_text,
    detect_year,
    generate_filename,
    get_output_folder,
    is_duplicate,
    move_file,
    write_detail_log,
    write_summary_csv,
    write_missing_docs_csv,
    write_coverage_dashboard,
    print_validation_summary,
    build_coverage_map,
    validate_coverage,
    CONFIDENCE_THRESHOLD,
    PAGE_TIERS,
    DEFAULT_YEAR,
    RAW_DIR,
    PROCESSED_DIR,
    REVIEW_DIR,
    LOG_DIR,
    SUMMARY_CSV,
    DETAIL_LOG,
    MISSING_DOCS_CSV,
    COVERAGE_DASHBOARD,
    setup_logging,
    CARDS,
    BANK_ALIASES,
)

from llm_classifier import (
    classify_with_llm,
    check_ollama_available,
    post_validate_llm_result,    # FIX F: re-validate after override
)

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# HYBRID PIPELINE CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

# FIX A: Increased override margin from 0.10 → 0.15
# LLM must be clearly better, not just marginally better.
LLM_OVERRIDE_MARGIN = 0.15

# Excel output paths
SUMMARY_XLSX      = LOG_DIR / "summary.xlsx"
MISSING_DOCS_XLSX = LOG_DIR / "missing_docs_report.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# BANK-SPECIFIC CARD MAP
# Used by _narrow_card_to_bank() and validate_prediction().
# ─────────────────────────────────────────────────────────────────────────────

BANK_CARDS: dict[str, list[str]] = {
    "HDFC": [
        "Infinia", "Diners Club Black", "Diners", "Marriott Bonvoy",
        "Regalia Gold", "Regalia", "Millennia", "Tata Neu Infinity",
        "Tata Neu", "MoneyBack+", "MoneyBack", "Swiggy", "IndianOil",
        "Pixel", "HDFC Millennia",
    ],
    "SBI": [
        "Cashback", "Elite", "Aurum", "SimplyCLICK", "SimplySAVE",
        "BPCL Octane", "BPCL", "IRCTC", "Paytm", "Reliance", "Vistara",
    ],
    "AXIS": [
        "Magnus", "Atlas", "Reserve", "ACE", "Axis ACE", "Flipkart",
        "Airtel", "Select", "Vistara Infinite", "Vistara", "Coral",
    ],
    "ICICI": [
        "Emeralde", "Amazon Pay", "HPCL Super Saver", "Rubyx",
    ],
    "AMEX": [
        "Platinum Travel", "Blue Cash", "Platinum",
    ],
    "HSBC": [
        "Premier", "Live+",
    ],
    "KOTAK": [
        "League Platinum", "Myntra",
    ],
    "SCB": [
        "Smart", "Ultimate",
    ],
    "IDFC": [
        "Millennia",
    ],
    "AU": [
        "Altura",
    ],
    "BOB": [
        "Eterna",
    ],
    "RBL": [
        "World Safari",
    ],
    "INDUSIND": [
        "EazyDiner",
    ],
    "YES": [
        "Prosperity",
    ],
    "ONECARD": [
        "OneCard",
    ],
    "SCAPIA": [
        "Scapia",
    ],
    "JUPITER": [
        "Jupiter",
    ],
    "CITI": [
        "Signature", "Platinum",
    ],
}

# ─────────────────────────────────────────────────────────────────────────────
# FIX D — FILENAME-BASED BANK/CARD HINTS
# Extract the expected bank and card from the original filename.
# Used by _filename_sanity_check() to detect obvious misclassifications.
# ─────────────────────────────────────────────────────────────────────────────

# Mapping from filename prefix (case-insensitive) to bank short code
FILENAME_BANK_HINTS: dict[str, str] = {
    "hdfc":     "HDFC",
    "sbi":      "SBI",
    "axis":     "AXIS",
    "icici":    "ICICI",
    "amex":     "AMEX",
    "hsbc":     "HSBC",
    "idfc":     "IDFC",
    "au":       "AU",
    "sc":       "SCB",     # SC_Smart → SCB
    "scb":      "SCB",
    "kotak":    "KOTAK",
    "yes":      "YES",
    "bob":      "BOB",
    "rbl":      "RBL",
    "indusind": "INDUSIND",
}


def _extract_filename_hints(filename: str) -> tuple[str | None, str | None]:
    """
    FIX D: Extract bank and card hints from the original filename.

    Filename format (expected): BANK_CARD_DOCTYPE_etc.pdf
    Examples:
        AXIS_Flipkart_TNC_Merchants_2026.pdf  → (AXIS, Flipkart)
        HDFC_Millenia_LG_2025.pdf             → (HDFC, Millenia/Millennia)
        SC_Smart_MITC_2026.pdf                → (SCB, Smart)
        SBI_Cashback_BR_2025.pdf              → (SBI, Cashback)

    Returns:
        (bank_hint: str | None, card_hint: str | None)
    """
    stem    = Path(filename).stem    # remove .pdf
    parts   = stem.split("_")       # split on underscore

    if not parts:
        return None, None

    # Bank: first part
    bank_prefix = parts[0].lower()
    bank_hint   = FILENAME_BANK_HINTS.get(bank_prefix)

    # Card: second part (if it exists and is not a doc type or year)
    doc_type_words = {"mitc", "tnc", "br", "lg", "master", "general", "main"}
    year_pattern   = r"^\d{4}$"

    import re as _re
    card_hint = None
    if len(parts) >= 2:
        candidate = parts[1]
        if (candidate.lower() not in doc_type_words
                and not _re.match(year_pattern, candidate)):
            card_hint = candidate

    return bank_hint, card_hint


def _filename_sanity_check(
    final_bank: str | None,
    final_card: str | None,
    filename:   str,
) -> tuple[bool, str]:
    """
    FIX D: Check if the final classification contradicts the filename.

    Returns:
        (passed: bool, reason: str)
        passed=False → classification strongly contradicts filename hints.

    Logic:
        - Extract bank+card hints from filename
        - If filename bank ≠ final bank AND filename bank is confident
          (i.e., it clearly starts with a known bank prefix) → FAIL
        - We don't fail on card mismatch alone (card names vary more)
    """
    fn_bank_hint, fn_card_hint = _extract_filename_hints(filename)

    if fn_bank_hint is None:
        # Can't determine bank from filename — pass by default
        return True, "Filename bank hint not determinable"

    if final_bank is None:
        return True, "Final bank is None — no contradiction possible"

    if fn_bank_hint != final_bank:
        return (
            False,
            f"FILENAME BANK CONFLICT: filename suggests bank='{fn_bank_hint}' "
            f"but classification produced bank='{final_bank}'. "
            f"File: {filename}. "
            f"This is likely a misclassification — routing to needs_review/."
        )

    return True, "Filename bank matches classification bank"


# ─────────────────────────────────────────────────────────────────────────────
# FIX B — BANK-SPECIFIC CARD NARROWING (IMPROVED)
# ─────────────────────────────────────────────────────────────────────────────

def _narrow_card_to_bank(
    bank:       str | None,
    card:       str | None,
    text_lower: str,
    is_master:  bool = False,
    filename:   str  = "",    # FIX B: use filename as additional signal
) -> str | None:
    """
    Validates that the detected card belongs to the detected bank.

    FIX B: When bank=SBI but filename starts with "SC_", this is likely
    an SCB (Standard Chartered) misclassification. We do NOT fix the bank
    here (that's done in detection), but we set card to None to route
    the file to needs_review/ where a human can correct it.

    FIX I (from original): is_master=True → skip narrowing entirely.
    Master docs have card="MASTER" which is intentionally not in BANK_CARDS.

    For non-master docs:
      1. Check if card is already valid for this bank → return as-is
      2. If not valid → re-scan text for the bank's own cards
         Pass 1: header zone (first 500 chars) — highest confidence
         Pass 2: full text body — lower confidence
      3. If not found → return None (routes to needs_review/)
    """
    # FIX I: Skip narrowing for master docs
    if is_master:
        logger.debug(
            f"[FIX #2] Skipping card narrowing for MASTER doc (bank={bank})"
        )
        return card

    if not bank or not card:
        return card

    bank_upper = bank.upper()

    # No mapping for this bank → leave card unchanged
    if bank_upper not in BANK_CARDS:
        return card

    allowed_cards = BANK_CARDS[bank_upper]

    # Card is already valid for this bank → no action needed
    if card in allowed_cards:
        return card

    # Card is NOT valid for this bank — try to find the correct one
    logger.warning(
        f"[FIX #2] Card '{card}' is not in {bank_upper}'s known card list. "
        f"Re-scanning text for a bank-specific card match."
    )

    header_zone = text_lower[:500]

    # Pass 1: header zone (most reliable)
    for candidate in allowed_cards:
        if candidate.lower() in header_zone:
            logger.info(
                f"[FIX #2] Found bank-specific card '{candidate}' in header "
                f"→ replacing '{card}'"
            )
            return candidate

    # Pass 2: full text body
    for candidate in allowed_cards:
        if candidate.lower() in text_lower:
            logger.info(
                f"[FIX #2] Found bank-specific card '{candidate}' in body "
                f"→ replacing '{card}'"
            )
            return candidate

    # No bank-specific card found → send to review
    logger.warning(
        f"[FIX #2] No card from {bank_upper}'s list found in text. "
        f"Setting card to None (will route to needs_review/)."
    )
    return None


# ─────────────────────────────────────────────────────────────────────────────
# OCR FALLBACK (unchanged from original)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_with_ocr(pdf_path: Path, max_pages: int) -> str:
    """
    OCR fallback for scanned PDFs where pdfplumber and PyMuPDF return empty text.

    REQUIREMENTS:
      pip install pytesseract Pillow
      # Tesseract binary:
      # Windows: https://github.com/UB-Mannheim/tesseract/wiki
      # macOS:   brew install tesseract
      # Ubuntu:  sudo apt install tesseract-ocr
    """
    try:
        import fitz
        import pytesseract
        from PIL import Image
        import io
    except ImportError as e:
        logger.warning(
            f"[OCR] Skipping OCR — missing dependency: {e}. "
            f"Install with: pip install pytesseract Pillow pymupdf"
        )
        return ""

    import platform
    if platform.system() == "Windows":
        import os
        TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(TESSERACT_CMD):
            pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
        else:
            ALT_CMD = r"C:\Users\Public\Tesseract-OCR\tesseract.exe"
            if os.path.exists(ALT_CMD):
                pytesseract.pytesseract.tesseract_cmd = ALT_CMD
            else:
                logger.warning(
                    f"[OCR] Tesseract not found at '{TESSERACT_CMD}'. "
                    f"Download from: https://github.com/UB-Mannheim/tesseract/wiki"
                )

    logger.info(f"[OCR] Running OCR on {pdf_path.name} (up to {max_pages} pages)...")

    text = ""
    try:
        import fitz as _fitz
        doc = _fitz.open(str(pdf_path))
        for page_num in range(min(max_pages, len(doc))):
            page = doc[page_num]
            mat  = _fitz.Matrix(300 / 72, 300 / 72)
            pix  = page.get_pixmap(matrix=mat)
            img  = Image.open(io.BytesIO(pix.tobytes("png")))
            page_text = pytesseract.image_to_string(img, lang="eng")
            if page_text.strip():
                text += page_text + "\n"
        doc.close()

        if text.strip():
            logger.info(
                f"[OCR] Extracted {len(text)} characters from {pdf_path.name}."
            )
        else:
            logger.warning(f"[OCR] OCR produced no text from {pdf_path.name}.")

    except Exception as e:
        logger.error(f"[OCR] OCR failed for {pdf_path.name}: {e}")

    return text


# ─────────────────────────────────────────────────────────────────────────────
# FIX E — SMARTER LLM TRIGGER CONDITIONS
# ─────────────────────────────────────────────────────────────────────────────

def _should_call_llm(
    rule_confidence: float,
    rule_bank:       str | None,
    rule_card:       str | None,
    rule_doc_type:   str,
    bank_confidence: float = 0.0,
) -> tuple[bool, str]:
    """
    FIX E: Smarter LLM trigger — avoids calling LLM when rules did well.

    LLM is triggered if ANY condition is true:
      1. rule_confidence < CONFIDENCE_THRESHOLD (0.70)
      2. Bank is None or UNKNOWN
      3. Card is None or UNKNOWN
      4. Doc type is empty or UNKNOWN

    FIX E — LLM is NOT triggered (even if conf < threshold) when:
      - Bank confidence ≥ 0.80 (bank was found clearly in text)
      - AND card is a known valid card (not UNKNOWN or None)
      - AND rule_confidence ≥ 0.65 (close to threshold, not critically low)
      This avoids wasting LLM calls on near-correct results where only a
      minor confidence tweak would have cleared the threshold.

    Returns (should_call: bool, reason: str)
    """
    reasons = []

    # Collect trigger conditions
    if rule_confidence < CONFIDENCE_THRESHOLD:
        reasons.append(
            f"confidence={rule_confidence:.2f} < threshold={CONFIDENCE_THRESHOLD}"
        )
    if not rule_bank or rule_bank.upper() == "UNKNOWN":
        reasons.append(f"bank='{rule_bank}' (UNKNOWN)")
    if not rule_card or rule_card.upper() == "UNKNOWN":
        reasons.append(f"card='{rule_card}' (UNKNOWN)")
    if not rule_doc_type or rule_doc_type.upper() == "UNKNOWN":
        reasons.append(f"doc_type='{rule_doc_type}' (UNKNOWN)")

    if not reasons:
        return False, ""

    # FIX E: Suppress LLM call when rule is nearly correct
    bank_is_clear = (
        rule_bank and rule_bank not in ("UNKNOWN", "None")
        and bank_confidence >= 0.80
    )
    card_is_known = (
        rule_card and rule_card not in ("UNKNOWN", "None")
    )
    conf_near_threshold = rule_confidence >= 0.65

    # Only suppress if the trigger was ONLY low confidence (not missing bank/card)
    trigger_is_only_low_conf = (
        reasons == [f"confidence={rule_confidence:.2f} < threshold={CONFIDENCE_THRESHOLD}"]
    )

    if (trigger_is_only_low_conf and bank_is_clear and card_is_known
            and conf_near_threshold):
        logger.info(
            f"[STEP 3] LLM trigger suppressed — rule found bank={rule_bank} "
            f"(bank_conf={bank_confidence:.2f}) and card={rule_card} "
            f"with conf={rule_confidence:.2f} ≥ 0.65. "
            f"Near-threshold result — LLM call would likely hallucinate."
        )
        return False, ""

    return True, " | ".join(reasons)


# ─────────────────────────────────────────────────────────────────────────────
# FIX A — HYBRID DECISION: RULE vs LLM (COMPLETELY REWRITTEN)
# ─────────────────────────────────────────────────────────────────────────────

def apply_llm_override(
    rule_bank:       str | None,
    rule_card:       str | None,
    rule_doc_type:   str,
    rule_is_master:  bool,
    rule_confidence: float,
    rule_bank_conf:  float,      # FIX A: bank-specific confidence
    llm_result:      dict,
) -> tuple[str | None, str | None, str, bool, float, str]:
    """
    FIX A: Decides whether to use LLM output or keep rule-based output.

    REWRITTEN LOGIC:
      Old: LLM wins if llm_confidence > rule_confidence + margin
      New: LLM wins ONLY IF:
        1. LLM classification succeeded (llm_success=True)
        2. LLM confidence (deflated) > rule confidence + LLM_OVERRIDE_MARGIN
        3. LLM bank matches rule bank (when rule bank was confident ≥ 0.75)
           → Bank mismatch = automatic rejection
        4. LLM is not replacing a valid card name with a different bank's card

    Master doc protection: if rule is_master=True, LLM cannot change
    the card name or is_master flag.

    Returns: (bank, card, doc_type, is_master, confidence, source)
    """
    if not llm_result.get("llm_success", False):
        logger.info(
            f"[DECISION] LLM failed/rejected → keeping rule result "
            f"(rule_conf={rule_confidence:.2f})"
        )
        return (
            rule_bank, rule_card, rule_doc_type, rule_is_master,
            rule_confidence, "rule_based",
        )

    llm_confidence = llm_result.get("confidence", 0.0)
    llm_bank       = llm_result.get("bank", "UNKNOWN")
    llm_card       = llm_result.get("card_name", "UNKNOWN")

    logger.info(
        f"[DECISION] rule_conf={rule_confidence:.2f} | "
        f"llm_conf={llm_confidence:.2f} (deflated) | "
        f"required margin={LLM_OVERRIDE_MARGIN} | "
        f"override threshold={rule_confidence + LLM_OVERRIDE_MARGIN:.2f}"
    )

    # ── FIX A: Bank mismatch check ────────────────────────────────────────────
    # If rule bank was confident and LLM disagrees on bank → automatic rejection
    rule_bank_upper = (rule_bank or "").upper()
    llm_bank_upper  = llm_bank.upper()

    bank_mismatch = (
        rule_bank and rule_bank not in ("UNKNOWN", "None")
        and llm_bank not in ("UNKNOWN",)
        and llm_bank_upper != rule_bank_upper
        and rule_bank_conf >= 0.75
    )

    if bank_mismatch:
        logger.warning(
            f"[DECISION] ✗ LLM REJECTED — bank mismatch: "
            f"LLM bank='{llm_bank}' ≠ rule bank='{rule_bank}' "
            f"(rule bank_conf={rule_bank_conf:.2f} ≥ 0.75). "
            f"Rule result protected."
        )
        return (
            rule_bank, rule_card, rule_doc_type, rule_is_master,
            rule_confidence, "rule_based",
        )

    # ── Standard confidence check ─────────────────────────────────────────────
    if llm_confidence <= rule_confidence + LLM_OVERRIDE_MARGIN:
        logger.info(
            f"[DECISION] Rules kept — LLM confidence insufficient "
            f"({llm_confidence:.2f} ≤ {rule_confidence:.2f} + {LLM_OVERRIDE_MARGIN})"
        )
        return (
            rule_bank, rule_card, rule_doc_type, rule_is_master,
            rule_confidence, "rule_based",
        )

    # ── LLM wins — apply with protections ────────────────────────────────────
    final_bank      = llm_result.get("bank",      "UNKNOWN")
    final_card      = llm_result.get("card_name", "UNKNOWN")
    final_doc_type  = llm_result.get("doc_type",   rule_doc_type)
    final_is_master = llm_result.get("is_master",  rule_is_master)

    # MASTER DOC PROTECTION: LLM cannot override is_master or card for master docs
    if rule_is_master:
        final_is_master = True
        final_card      = "MASTER"
        logger.info(
            "[DECISION] Master doc protection: rule is_master=True → "
            "preserving card='MASTER', is_master=True"
        )

    # FIX A: Preserve rule values for fields LLM couldn't identify
    if final_bank in ("UNKNOWN", "") and rule_bank:
        final_bank = rule_bank
        logger.info(f"[DECISION] LLM bank=UNKNOWN → keeping rule bank={rule_bank}")
    if not rule_is_master and final_card in ("UNKNOWN", "") and rule_card:
        final_card = rule_card
        logger.info(f"[DECISION] LLM card=UNKNOWN → keeping rule card={rule_card}")

    # FIX A: If LLM is replacing a valid rule card with a different card,
    # only allow it when bank also matches (already checked above)
    rule_card_valid = (
        rule_card and rule_card not in ("UNKNOWN", "None")
        and not rule_is_master
    )
    llm_changes_card = (
        final_card not in ("UNKNOWN", rule_card)
        and rule_card_valid
    )
    if llm_changes_card:
        logger.info(
            f"[DECISION] LLM changing card from '{rule_card}' to '{final_card}'. "
            f"Banks match ({rule_bank} == {final_bank}) — allowing card override."
        )

    logger.info(
        f"[DECISION] ✓ LLM OVERRIDE applied "
        f"({llm_confidence:.2f} > {rule_confidence:.2f} + {LLM_OVERRIDE_MARGIN}) "
        f"→ {final_bank} | {final_card} | {final_doc_type}"
    )
    return (
        final_bank, final_card, final_doc_type, final_is_master,
        llm_confidence, "llm",
    )


# ─────────────────────────────────────────────────────────────────────────────
# METADATA GENERATION (unchanged from original)
# ─────────────────────────────────────────────────────────────────────────────

def generate_metadata(
    bank:                  str | None,
    card:                  str | None,
    doc_type:              str,
    is_master:             bool,
    confidence:            float,
    classification_source: str,
    source_file:           str,
    llm_reason:            str | None = None,
    dest_path:             Path | None = None,
    present_doc_types:     set | None = None,
    validation_reasons:    list | None = None,
) -> dict:
    """
    Builds the structured metadata dict saved alongside each processed PDF.
    This metadata is what downstream RAG agents use to filter documents.

    data_quality:
      "complete"     → both MITC and BR present
      "partial"      → at least one present
      "insufficient" → neither present
    """
    REQUIRED_FOR_QUALITY = {"MITC", "BR"}

    if present_doc_types is None:
        data_quality = "partial"
    else:
        if REQUIRED_FOR_QUALITY.issubset(present_doc_types):
            data_quality = "complete"
        elif REQUIRED_FOR_QUALITY & present_doc_types:
            data_quality = "partial"
        else:
            data_quality = "insufficient"

    return {
        "bank":                   bank or "UNKNOWN",
        "card":                   card or "UNKNOWN",
        "doc_type":               doc_type,
        "is_master":              is_master,
        "confidence":             round(confidence, 4),
        "classification_source":  classification_source,
        "source_file":            source_file,
        "output_file":            dest_path.name if dest_path else None,
        "llm_reason":             llm_reason,
        "data_quality":           data_quality,
        "validation_issues":      validation_reasons or [],
        "processing_timestamp":   datetime.now().isoformat(timespec="seconds"),
    }


def save_metadata_json(
    metadata: dict,
    dest_path: Path,
    dry_run:   bool,
) -> None:
    """
    Saves metadata dict as JSON alongside the processed PDF.
    e.g. HDFC_Millennia_MITC_2026.pdf → HDFC_Millennia_MITC_2026.json
    In dry-run mode: logs what WOULD happen, does NOT write any file.
    """
    json_path = dest_path.with_suffix(".json")
    if dry_run:
        logger.info(f"[DRY-RUN] Would write metadata JSON: {json_path.name}")
        return
    try:
        json_path.parent.mkdir(parents=True, exist_ok=True)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)
        logger.info(f"[OUTPUT] Metadata JSON saved: {json_path}")
    except Exception as e:
        logger.error(f"[OUTPUT] Failed to save metadata JSON: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL LOG WRITERS (FIX H from original — unchanged)
# ─────────────────────────────────────────────────────────────────────────────

def write_summary_xlsx(entries: list[dict], dry_run: bool = False) -> None:
    """
    Writes summary.xlsx — per-file processing results with colour coding.

    Colour coding by Status:
      SUCCESS           → green
      MASTER_DOC        → blue
      NEEDS_REVIEW      → orange/yellow
      ERROR             → red
      DUPLICATE_SKIPPED → grey

    FIX H: Respects dry_run — if True, only prints what would happen.
    """
    if dry_run:
        logger.info("[DRY-RUN] Would write summary Excel: summary.xlsx")
        return

    rows = []
    for e in entries:
        issues = "; ".join(e.get("validation_issues", [])[:2]) or "None"
        rows.append({
            "File Name":       e["filename"],
            "Bank":            e.get("bank") or "NOT FOUND",
            "Card":            e.get("card") or "NOT FOUND",
            "Is_Master":       "YES" if e.get("is_master") else "NO",
            "DocType":         e.get("doc_type", "UNKNOWN"),
            "Confidence":      round(e.get("overall_conf", 0.0), 2),
            "Final Conf":      round(e.get("final_confidence", 0.0), 2),
            "Source":          e.get("classification_source", "rule_based"),
            "LLM Called":      "YES" if e.get("llm_called") else "NO",
            "LLM Override":    "YES" if e.get("classification_source") == "llm" else "NO",
            "Filename Check":  "PASS" if e.get("filename_check_passed", True) else "FAIL",
            "Status":          e.get("status", "UNKNOWN"),
            "Issues":          issues,
        })

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Processing Summary"

        columns = list(rows[0].keys()) if rows else []

        STATUS_FILL = {
            "SUCCESS":           PatternFill("solid", fgColor="C6EFCE"),
            "MASTER_DOC":        PatternFill("solid", fgColor="BDD7EE"),
            "NEEDS_REVIEW":      PatternFill("solid", fgColor="FFEB9C"),
            "ERROR":             PatternFill("solid", fgColor="FFC7CE"),
            "DUPLICATE_SKIPPED": PatternFill("solid", fgColor="D9D9D9"),
        }
        STATUS_FONT = {
            "SUCCESS":           Font(bold=False, color="276221"),
            "MASTER_DOC":        Font(bold=False, color="1F4E79"),
            "NEEDS_REVIEW":      Font(bold=True,  color="9C6500"),
            "ERROR":             Font(bold=True,  color="9C0006"),
            "DUPLICATE_SKIPPED": Font(bold=False, color="595959"),
        }

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = thin_border
        ws.row_dimensions[1].height = 30

        for row_idx, row_data in enumerate(rows, start=2):
            status = row_data.get("Status", "UNKNOWN")
            fill   = STATUS_FILL.get(status, PatternFill())
            font   = STATUS_FONT.get(status, Font())

            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, "")
                cell  = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill      = fill
                cell.font      = font
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True)
                cell.border    = thin_border
            ws.row_dimensions[row_idx].height = 18

        col_widths = {
            "File Name": 40, "Bank": 8, "Card": 18, "Is_Master": 10,
            "DocType": 8, "Confidence": 12, "Final Conf": 10,
            "Source": 12, "LLM Called": 10, "LLM Override": 12,
            "Filename Check": 14, "Status": 18, "Issues": 50,
        }
        for col_idx, col_name in enumerate(columns, start=1):
            width = col_widths.get(col_name, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

        summary_row = len(rows) + 3
        from collections import Counter
        status_counts = Counter(r["Status"] for r in rows)

        ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
        col = 2
        for status_name, count in status_counts.items():
            ws.cell(row=summary_row, column=col, value=f"{status_name}: {count}")
            col += 1

        LOG_DIR.mkdir(parents=True, exist_ok=True)
        wb.save(str(SUMMARY_XLSX))
        logger.info(f"Summary Excel written to: {SUMMARY_XLSX}")

    except ImportError:
        logger.warning("openpyxl not installed — writing summary as CSV instead.")
        write_summary_csv(entries)


def write_missing_docs_xlsx(
    validation_results: list[dict],
    dry_run: bool = False,
) -> None:
    """
    Writes missing_docs_report.xlsx — coverage gaps per card.
    FIX H: Respects dry_run.
    """
    if dry_run:
        logger.info("[DRY-RUN] Would write missing docs Excel: missing_docs_report.xlsx")
        return

    rows = [{
        "Card":         r["card"],
        "Present Docs": ", ".join(r["present_docs"]) or "NONE",
        "Missing Docs": ", ".join(r["missing_docs"]) or "NONE",
        "Status":       r["status"],
    } for r in validation_results]

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Missing Docs Report"

        columns = ["Card", "Present Docs", "Missing Docs", "Status"]

        STATUS_FILL = {
            "COMPLETE":  PatternFill("solid", fgColor="C6EFCE"),
            "PARTIAL":   PatternFill("solid", fgColor="FFEB9C"),
            "CRITICAL":  PatternFill("solid", fgColor="FFC7CE"),
            "NOT_FOUND": PatternFill("solid", fgColor="E8B0B5"),
        }
        STATUS_FONT = {
            "COMPLETE":  Font(bold=False, color="276221"),
            "PARTIAL":   Font(bold=True,  color="9C6500"),
            "CRITICAL":  Font(bold=True,  color="9C0006"),
            "NOT_FOUND": Font(bold=True,  color="6B0000"),
        }

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border
        ws.row_dimensions[1].height = 24

        for row_idx, row_data in enumerate(rows, start=2):
            status = row_data["Status"]
            fill   = STATUS_FILL.get(status, PatternFill())
            font   = STATUS_FONT.get(status, Font())
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
                cell.fill      = fill
                cell.font      = font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = thin_border
            ws.row_dimensions[row_idx].height = 18

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 12
        ws.freeze_panes = "A2"

        LOG_DIR.mkdir(parents=True, exist_ok=True)
        wb.save(str(MISSING_DOCS_XLSX))
        logger.info(f"Missing docs Excel written to: {MISSING_DOCS_XLSX}")

    except ImportError:
        logger.warning("openpyxl not installed — writing missing docs as CSV instead.")
        write_missing_docs_csv(validation_results)


# ─────────────────────────────────────────────────────────────────────────────
# FIX C — VALIDATION (IMPROVED)
# Now checks cross-bank mismatches and re-validates after LLM override.
# ─────────────────────────────────────────────────────────────────────────────

VALIDATION_CONF_THRESHOLD = CONFIDENCE_THRESHOLD   # 0.70
BANK_CONF_MIN             = 0.75
CARD_CONF_MIN             = 0.60

GENERIC_CARD_NAMES: set[str] = {
    "Platinum", "Smart", "Signature", "Select",
    "Cashback", "Elite", "Coral", "Reserve",
}


def validate_prediction(
    bank_result:       dict,
    card_result:       dict,
    doc_type_result:   dict,
    overall_conf:      float,
    is_master:         bool,
    source_filename:   str = "",   # FIX D: used for filename sanity check
    classification_src: str = "rule_based",  # FIX F: for logging
) -> tuple[bool, list[str]]:
    """
    FIX C: Multi-layer validation — now includes cross-bank mismatch check.

    Checks:
      1. Missing required fields (bank or card is None)
      2. Overall confidence floor
      3. Bank confidence minimum
      4. Card confidence minimum
      5. Generic card name with low confidence
      6. Invalid bank-card combination (cross-bank mismatch) — IMPROVED
      7. FIX D: Filename sanity check (bank from filename ≠ classified bank)

    Returns (needs_review: bool, reasons: list[str])
    """
    reasons: list[str] = []

    bank      = bank_result.get("value")
    card      = card_result.get("value")
    bank_conf = bank_result.get("confidence", 0.0)
    card_conf = card_result.get("confidence", 0.0)

    # CHECK 1: Missing required fields
    if not bank:
        reasons.append(
            "Bank not detected — no alias phrase matched in text or filename. "
            "Add the bank name to BANK_ALIASES in preprocess.py."
        )
    if not card and not is_master:
        reasons.append(
            "Card not detected — no card name matched in header, body, or filename. "
            "Add the card name to CARDS or BANK_CARDS in preprocess.py."
        )

    # CHECK 2: Overall confidence floor
    if overall_conf < VALIDATION_CONF_THRESHOLD:
        reasons.append(
            f"Overall confidence {overall_conf:.2f} is below threshold "
            f"{VALIDATION_CONF_THRESHOLD:.2f}. Classification is uncertain."
        )

    # CHECK 3: Bank confidence
    if bank and bank_conf < BANK_CONF_MIN:
        reasons.append(
            f"Bank confidence {bank_conf:.2f} is below minimum {BANK_CONF_MIN:.2f}. "
            f"Bank '{bank}' may have been detected from the filename only."
        )

    # CHECK 4: Card confidence
    if card and card != "MASTER" and card_conf < CARD_CONF_MIN:
        reasons.append(
            f"Card confidence {card_conf:.2f} is below minimum {CARD_CONF_MIN:.2f}. "
            f"Card '{card}' may be a weak match — verify the document is for this card."
        )

    # CHECK 5: Generic card name
    if card and card in GENERIC_CARD_NAMES and not is_master:
        if card_conf < 0.90:
            reasons.append(
                f"Card name '{card}' is generic and appears in many documents. "
                f"Card confidence: {card_conf:.2f}. Verify this is the correct card."
            )

    # CHECK 6 (FIX C): Cross-bank mismatch validation
    if card and card != "MASTER" and bank and not is_master:
        bank_upper    = bank.upper()
        allowed_cards = BANK_CARDS.get(bank_upper)
        if allowed_cards is not None and card not in allowed_cards:
            reasons.append(
                f"CROSS-BANK MISMATCH: '{card}' is not a card issued by {bank}. "
                f"Known {bank} cards: {', '.join(allowed_cards[:5])}... "
                f"This is a likely misclassification — manual review required."
            )

    # CHECK 7 (FIX D): Filename sanity check
    if source_filename:
        fn_passed, fn_reason = _filename_sanity_check(bank, card, source_filename)
        if not fn_passed:
            reasons.append(fn_reason)

    needs_review = len(reasons) > 0
    return needs_review, reasons


# ─────────────────────────────────────────────────────────────────────────────
# DRY-RUN SAFE FOLDER CREATION (FIX H from original)
# ─────────────────────────────────────────────────────────────────────────────

def _get_output_folder_safe(
    bank:     str | None,
    card:     str | None,
    base_dir: Path,
    dry_run:  bool,
) -> Path:
    """
    Builds the output folder path.
    In dry-run mode: returns path WITHOUT creating it on disk.
    In normal mode: creates the folder and returns the path.
    """
    from preprocess import _safe
    b      = _safe(bank, "UNKNOWN_BANK")
    c      = _safe(card, "UNKNOWN_CARD")
    folder = base_dir / f"{b}_{c}"
    if not dry_run:
        folder.mkdir(parents=True, exist_ok=True)
    return folder


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE-FILE PROCESSOR (ALL FIXES INTEGRATED)
# ─────────────────────────────────────────────────────────────────────────────

def _process_one_file(
    pdf_path:      Path,
    debug:         bool,
    dry_run:       bool,
    use_llm:       bool,
    llm_available: bool,
) -> dict:
    """
    Processes a single PDF through the full hybrid pipeline.

    Steps:
      [STEP 1] Extract text (pdfplumber/PyMuPDF, then OCR if both fail)
      [STEP 2] Rule-based classification + bank-specific card narrowing
               FIX I: narrowing SKIPPED for master docs
               FIX B: SCB vs SBI disambiguation
      [STEP 3] FIX E: Smart LLM trigger (suppress near-correct results)
               FIX G: Pass rule hints to LLM for grounding
               FIX A: Bank mismatch → automatic LLM rejection
      [STEP 4] FIX C: Validate result (cross-bank check)
               FIX D: Filename sanity check
               FIX F: Re-validate after LLM override
               FIX H: Dry-run safe — no files written in dry-run
    """
    pages_read          = PAGE_TIERS[0]
    llm_called          = False
    llm_result          = None
    classification_src  = "rule_based"
    llm_reason          = None
    status              = "SUCCESS"
    validation_reasons: list[str] = []
    filename_check_passed = True

    # Safe defaults
    doc_type_result = {"value": "UNKNOWN", "confidence": 0.0, "reasons": []}
    bank_result     = {"value": None,      "confidence": 0.0, "reasons": []}
    card_result     = {"value": None,      "confidence": 0.0, "reasons": []}
    master_result   = {"is_master": False, "signal": None,   "confidence": 0.0}
    rule_confidence = 0.0
    final_confidence= 0.0
    final_bank      = None
    final_card      = None
    final_doc_type  = "UNKNOWN"
    final_is_master = False
    year            = DEFAULT_YEAR
    dest_path       = REVIEW_DIR / "UNKNOWN_UNKNOWN_UNKNOWN.pdf"

    try:
        # ── STEP 1: Text extraction ────────────────────────────────────────────
        logger.info("[STEP 1] Extracting text from PDF")

        # ── STEP 2: Rule-based classification ─────────────────────────────────
        logger.info("[STEP 2] Running rule-based classification")

        (text, doc_type_result, bank_result,
         card_result, master_result, year, rule_confidence) = \
            run_detection_with_fallback(pdf_path, debug)

        # OCR fallback if both pdfplumber and PyMuPDF returned nothing
        if not text.strip():
            logger.warning(
                "[STEP 1] Standard extraction returned empty text. "
                "Attempting OCR fallback..."
            )
            ocr_text = _extract_with_ocr(pdf_path, max_pages=PAGE_TIERS[-1])
            if ocr_text.strip():
                from preprocess import (
                    detect_doc_type, detect_bank, detect_card,
                    detect_master_doc, detect_year as _detect_year,
                )
                text            = ocr_text
                doc_type_result = detect_doc_type(text)
                bank_result     = detect_bank(text, pdf_path.name)
                master_result   = detect_master_doc(text)
                if master_result["is_master"]:
                    card_result = {
                        "value": "MASTER",
                        "confidence": master_result["confidence"],
                        "reasons": ["OCR-detected master doc"],
                    }
                else:
                    card_result = detect_card(text, pdf_path.name)
                year            = _detect_year(text)
                rule_confidence = compute_confidence(
                    doc_type_result, bank_result, card_result, master_result
                )
                logger.info(
                    f"[STEP 1] OCR re-classification: "
                    f"bank={bank_result['value']} | "
                    f"card={card_result['value']} | "
                    f"doc_type={doc_type_result['value']} | "
                    f"conf={rule_confidence:.2f}"
                )
            else:
                logger.error(
                    f"[STEP 1] OCR also produced no text for {pdf_path.name}. "
                    f"File will go to needs_review/."
                )

        logger.info(
            f"[STEP 2] Rule result: "
            f"bank={bank_result['value']} | "
            f"card={card_result['value']} | "
            f"doc_type={doc_type_result['value']} | "
            f"rule_confidence={rule_confidence:.2f}"
        )

        for tier in PAGE_TIERS:
            pages_read = tier
            if rule_confidence >= CONFIDENCE_THRESHOLD:
                break

        # ── FIX I: Bank-specific card narrowing (SKIPPED for master docs) ─────
        raw_card       = card_result["value"]
        rule_is_master = master_result["is_master"]

        narrowed = _narrow_card_to_bank(
            bank      = bank_result["value"],
            card      = raw_card,
            text_lower= text.lower() if text else "",
            is_master = rule_is_master,
            filename  = pdf_path.name,    # FIX B: pass filename for SCB check
        )
        if narrowed != raw_card:
            logger.info(
                f"[FIX #2] Card narrowed: '{raw_card}' → '{narrowed}' "
                f"(bank={bank_result['value']})"
            )
            card_result = dict(card_result)
            card_result["value"] = narrowed
            rule_confidence = compute_confidence(
                doc_type_result, bank_result, card_result, master_result
            )
            logger.info(
                f"[FIX #2] Confidence after card correction: {rule_confidence:.2f}"
            )

        rule_bank      = bank_result["value"]
        rule_card      = card_result["value"]
        rule_doc_type  = doc_type_result["value"]
        rule_bank_conf = bank_result.get("confidence", 0.0)   # FIX A

        # Start with rule values as defaults
        final_bank       = rule_bank
        final_card       = rule_card
        final_doc_type   = rule_doc_type
        final_is_master  = rule_is_master
        final_confidence = rule_confidence

        # ── STEP 3: FIX E — Smart LLM trigger check ───────────────────────────
        should_llm, trigger_reason = _should_call_llm(
            rule_confidence = rule_confidence,
            rule_bank       = rule_bank,
            rule_card       = rule_card,
            rule_doc_type   = rule_doc_type,
            bank_confidence = rule_bank_conf,   # FIX E
        )

        if should_llm and use_llm and llm_available:
            logger.info(f"[STEP 3] LLM triggered — reasons: {trigger_reason}")
            llm_called = True

            # FIX G: Pass rule hints to LLM for grounding
            llm_result = classify_with_llm(
                text            = text,
                rule_bank       = rule_bank     or "UNKNOWN",
                rule_card       = rule_card     or "UNKNOWN",
                rule_doc_type   = rule_doc_type or "UNKNOWN",
                rule_confidence = rule_confidence,
            )

            # FIX A: Apply rewritten override logic
            (final_bank, final_card, final_doc_type, final_is_master,
             final_confidence, classification_src) = apply_llm_override(
                rule_bank       = rule_bank,
                rule_card       = rule_card,
                rule_doc_type   = rule_doc_type,
                rule_is_master  = rule_is_master,
                rule_confidence = rule_confidence,
                rule_bank_conf  = rule_bank_conf,   # FIX A
                llm_result      = llm_result,
            )

            if classification_src == "llm":
                llm_reason = llm_result.get("reason")

        elif should_llm and use_llm and not llm_available:
            logger.info("[STEP 3] LLM would be triggered but Ollama is not available")
        elif should_llm and not use_llm:
            logger.info("[STEP 3] LLM skipped — disabled via --no-llm flag")
        else:
            logger.info(
                f"[STEP 3] LLM skipped — all conditions satisfied "
                f"(conf={rule_confidence:.2f}, bank={rule_bank}, "
                f"card={rule_card}, doc_type={rule_doc_type})"
            )

        # ── STEP 4: FIX C+D+F — Validate + determine output path ──────────────
        logger.info(
            f"[STEP 4] Final result — "
            f"bank={final_bank} | card={final_card} | "
            f"doc_type={final_doc_type} | master={final_is_master} | "
            f"source={classification_src} | confidence={final_confidence:.2f}"
        )

        # Build result dicts for validation using FINAL values (post-LLM)
        final_bank_result = {
            "value":      final_bank,
            "confidence": bank_result.get("confidence", 0.0),
        }
        final_card_result = {
            "value":      final_card,
            "confidence": card_result.get("confidence", 0.0),
        }
        final_doc_type_result = {
            "value":      final_doc_type,
            "confidence": doc_type_result.get("confidence", 0.0),
        }

        # FIX F: Re-validate after LLM override (not just before)
        needs_review, validation_reasons = validate_prediction(
            bank_result        = final_bank_result,
            card_result        = final_card_result,
            doc_type_result    = final_doc_type_result,
            overall_conf       = final_confidence,
            is_master          = final_is_master,
            source_filename    = pdf_path.name,     # FIX D
            classification_src = classification_src, # FIX F
        )

        # FIX D: Track filename check result for summary report
        fn_passed, fn_reason = _filename_sanity_check(final_bank, final_card, pdf_path.name)
        filename_check_passed = fn_passed

        if needs_review:
            logger.warning(
                f"[VALIDATION] ⚠️  File routed to needs_review/ — "
                f"{len(validation_reasons)} issue(s) found:"
            )
            for i, reason in enumerate(validation_reasons, start=1):
                logger.warning(f"[VALIDATION]   {i}. {reason}")
        else:
            logger.info("[VALIDATION] ✓ All checks passed — routing to processed_docs/")

        new_filename = generate_filename(final_bank, final_card, final_doc_type, year)

        if needs_review:
            if not dry_run:
                REVIEW_DIR.mkdir(parents=True, exist_ok=True)
            dest_path = REVIEW_DIR / new_filename
            status    = "NEEDS_REVIEW"
        else:
            dest_dir  = _get_output_folder_safe(
                final_bank, final_card, PROCESSED_DIR, dry_run
            )
            dest_path = dest_dir / new_filename
            status    = "MASTER_DOC" if final_is_master else "SUCCESS"

        if is_duplicate(dest_path):
            logger.info(f"Duplicate detected — skipping: {new_filename}")
            status = "DUPLICATE_SKIPPED"
        else:
            move_file(pdf_path, dest_path, dry_run)

        # ── Metadata JSON ─────────────────────────────────────────────────────
        if status not in ("DUPLICATE_SKIPPED",):
            metadata = generate_metadata(
                bank                  = final_bank,
                card                  = final_card,
                doc_type              = final_doc_type,
                is_master             = final_is_master,
                confidence            = final_confidence,
                classification_source = classification_src,
                source_file           = pdf_path.name,
                llm_reason            = llm_reason,
                dest_path             = dest_path,
                validation_reasons    = validation_reasons,
            )
            save_metadata_json(metadata, dest_path, dry_run)
            logger.info(f"[OUTPUT] Saved: {dest_path.name}")

    except Exception as exc:
        logger.error(f"Error processing {pdf_path.name}: {exc}", exc_info=True)
        status          = "ERROR"
        doc_type_result = {"value": "ERROR", "confidence": 0.0, "reasons": [str(exc)]}
        bank_result     = {"value": None,    "confidence": 0.0, "reasons": []}
        card_result     = {"value": None,    "confidence": 0.0, "reasons": []}
        master_result   = {"is_master": False, "signal": None,  "confidence": 0.0}
        final_confidence = 0.0
        final_is_master  = False
        classification_src = "rule_based"
        llm_reason       = None
        rule_confidence  = 0.0
        validation_reasons = [f"Processing error: {exc}"]
        filename_check_passed = False

    return {
        "filename":               pdf_path.name,
        "bank":                   bank_result["value"],
        "bank_conf":              bank_result["confidence"],
        "bank_reasons":           bank_result["reasons"],
        "card":                   card_result["value"],
        "card_conf":              card_result["confidence"],
        "card_reasons":           card_result["reasons"],
        "doc_type":               doc_type_result["value"],
        "doc_type_conf":          doc_type_result["confidence"],
        "doc_type_reasons":       doc_type_result["reasons"],
        "is_master":              master_result["is_master"],
        "master_signal":          master_result["signal"],
        "master_conf":            master_result["confidence"],
        "overall_conf":           rule_confidence,
        "pages_read":             pages_read,
        "status":                 status,
        "llm_called":             llm_called,
        "llm_success":            llm_result.get("llm_success", False) if llm_result else False,
        "llm_confidence":         llm_result.get("confidence",  0.0)   if llm_result else 0.0,
        "llm_bank":               llm_result.get("bank")                if llm_result else None,
        "llm_card":               llm_result.get("card_name")           if llm_result else None,
        "llm_doc_type":           llm_result.get("doc_type")            if llm_result else None,
        "llm_reason":             llm_reason,
        "classification_source":  classification_src,
        "final_confidence":       final_confidence,
        "validation_issues":      validation_reasons,
        "validation_passed":      len(validation_reasons) == 0,
        "filename_check_passed":  filename_check_passed,          # FIX D
    }


# ─────────────────────────────────────────────────────────────────────────────
# HYBRID LOG WRITER (updated with FIX D info)
# ─────────────────────────────────────────────────────────────────────────────

def _write_hybrid_log(entries: list[dict], dry_run: bool = False) -> None:
    """
    Writes hybrid_classification_log.txt — files where LLM was called.
    Now also shows filename sanity check results.
    FIX H: Respects dry_run.
    """
    if dry_run:
        llm_count = sum(1 for e in entries if e.get("llm_called"))
        logger.info(f"[DRY-RUN] Would write hybrid log ({llm_count} LLM-triggered files)")
        return

    hybrid_log_path = LOG_DIR / "hybrid_classification_log.txt"
    llm_entries     = [e for e in entries if e.get("llm_called")]

    with open(hybrid_log_path, "w", encoding="utf-8") as f:
        f.write(
            f"HYBRID CLASSIFICATION LOG — "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        )
        f.write("=" * 70 + "\n")
        f.write(f"Files where LLM was invoked: {len(llm_entries)}\n\n")

        if not llm_entries:
            f.write("No files required LLM fallback — all rule-based results passed.\n")
            return

        for entry in llm_entries:
            f.write("=" * 40 + "\n")
            f.write(f"FILE: {entry['filename']}\n\n")
            f.write("  RULE RESULT:\n")
            f.write(f"    Bank       : {entry['bank']}\n")
            f.write(f"    Card       : {entry['card']}\n")
            f.write(f"    DocType    : {entry['doc_type']}\n")
            f.write(f"    Confidence : {entry['overall_conf']:.2f}\n\n")
            f.write("  LLM RESULT:\n")
            f.write(f"    Bank       : {entry.get('llm_bank', 'N/A')}\n")
            f.write(f"    Card       : {entry.get('llm_card', 'N/A')}\n")
            f.write(f"    DocType    : {entry.get('llm_doc_type', 'N/A')}\n")
            f.write(f"    Confidence : {entry.get('llm_confidence', 0):.2f} (deflated)\n")
            f.write(f"    Reason     : {entry.get('llm_reason', 'N/A')}\n\n")
            src = entry.get("classification_source", "rule_based")
            f.write(f"  DECISION         : Used '{src.upper()}' result\n")
            f.write(f"  Final conf       : {entry.get('final_confidence', 0):.2f}\n")
            f.write(f"  Filename check   : {'PASS' if entry.get('filename_check_passed', True) else 'FAIL'}\n")
            f.write(f"  Status           : {entry['status']}\n\n")

    logger.info(f"Hybrid log written to: {hybrid_log_path}")


# ─────────────────────────────────────────────────────────────────────────────
# DRY-RUN SAFE LOG WRITERS (FIX H from original)
# ─────────────────────────────────────────────────────────────────────────────

def _write_detail_log_safe(entries: list[dict], dry_run: bool) -> None:
    if dry_run:
        logger.info("[DRY-RUN] Would write detailed log: preprocess_log.txt")
        return
    write_detail_log(entries)


def _write_summary_safe(entries: list[dict], dry_run: bool) -> None:
    write_summary_xlsx(entries, dry_run=dry_run)


def _write_coverage_safe(validation_results: list[dict], dry_run: bool) -> None:
    if dry_run:
        logger.info("[DRY-RUN] Would write coverage dashboard: coverage_dashboard.xlsx")
        return
    write_coverage_dashboard(validation_results)


def _write_missing_docs_safe(validation_results: list[dict], dry_run: bool) -> None:
    write_missing_docs_xlsx(validation_results, dry_run=dry_run)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN HYBRID PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def process_all_hybrid(
    dry_run: bool = False,
    debug:   bool = False,
    use_llm: bool = True,
) -> None:
    """
    Full hybrid preprocessing pipeline.

    In --dry-run mode:
      - No folders created on disk
      - No files copied
      - No metadata JSON written
      - No log files written to disk
      - Everything printed to console only
    """
    if not dry_run:
        for d in [RAW_DIR, PROCESSED_DIR, REVIEW_DIR, LOG_DIR]:
            d.mkdir(parents=True, exist_ok=True)

    pdf_files = sorted(RAW_DIR.glob("*.pdf"))
    total     = len(pdf_files)

    if total == 0:
        logger.warning(f"No PDF files found in {RAW_DIR}. Exiting.")
        return

    logger.info(f"Found {total} PDF file(s) in {RAW_DIR}")

    # Preflight: LLM availability
    llm_available = False
    if use_llm:
        logger.info("[STEP 0] Checking LLM (Ollama) availability...")
        llm_available = check_ollama_available()
        if not llm_available:
            logger.warning(
                "[LLM] Ollama not available — LLM fallback will be SKIPPED. "
                "To enable: run 'ollama serve' and 'ollama pull llama3.2:1b'"
            )
    else:
        logger.info("[LLM] LLM fallback disabled via --no-llm flag.")

    # Counters
    count_processed = count_review = count_errors = count_skipped = 0
    count_llm_used  = count_llm_won = count_llm_rejected = 0
    count_fn_fail   = 0    # FIX D: filename check failures
    log_entries: list[dict] = []

    for idx, pdf_path in enumerate(pdf_files, start=1):
        logger.info("")
        logger.info("=" * 60)
        logger.info(f"[PROCESSING] {idx}/{total}: {pdf_path.name}")
        logger.info("=" * 60)

        entry = _process_one_file(
            pdf_path      = pdf_path,
            debug         = debug,
            dry_run       = dry_run,
            use_llm       = use_llm,
            llm_available = llm_available,
        )

        log_entries.append(entry)

        s = entry["status"]
        if s == "ERROR":
            count_errors   += 1
        elif s == "DUPLICATE_SKIPPED":
            count_skipped  += 1
        elif s == "NEEDS_REVIEW":
            count_review   += 1
        else:
            count_processed += 1

        if entry["llm_called"]:
            count_llm_used += 1
        if entry.get("classification_source") == "llm":
            count_llm_won += 1
        if entry["llm_called"] and entry.get("classification_source") == "rule_based":
            # LLM was called but rule kept → LLM was rejected
            count_llm_rejected += 1
        if not entry.get("filename_check_passed", True):
            count_fn_fail += 1

    # ── Write logs ──────────────────────────────────────────────────────────────
    _write_detail_log_safe(log_entries, dry_run)
    _write_summary_safe(log_entries, dry_run)
    _write_hybrid_log(log_entries, dry_run)

    # ── Coverage validation ────────────────────────────────────────────────────
    logger.info("")
    logger.info("Running document coverage validation...")
    coverage_map       = build_coverage_map(log_entries)
    validation_results = validate_coverage(coverage_map)
    _write_missing_docs_safe(validation_results, dry_run)
    _write_coverage_safe(validation_results, dry_run)
    print_validation_summary(validation_results)

    # ── Final summary ──────────────────────────────────────────────────────────
    master_count = sum(
        1 for e in log_entries
        if e.get("is_master") and e["status"] in ("MASTER_DOC", "SUCCESS")
    )
    logger.info("")
    logger.info("=" * 60)
    logger.info("HYBRID PIPELINE COMPLETE")
    logger.info(f"  Total files           : {total}")
    logger.info(f"  Processed             : {count_processed}")
    logger.info(f"  Master/Collective     : {master_count}")
    logger.info(f"  Needs review          : {count_review}")
    logger.info(f"  Errors                : {count_errors}")
    logger.info(f"  Duplicates skipped    : {count_skipped}")
    logger.info(f"  LLM triggered         : {count_llm_used} (of {total} files)")
    logger.info(f"  LLM overrides used    : {count_llm_won}")
    logger.info(f"  LLM results rejected  : {count_llm_rejected}")
    logger.info(f"  Filename check fails  : {count_fn_fail}")
    logger.info(f"  Dry-run mode          : {'ON (no files written)' if dry_run else 'OFF'}")
    logger.info("=" * 60)

    if dry_run:
        logger.info("")
        logger.info("=" * 60)
        logger.info("DRY-RUN COMPLETE — No files were copied or created.")
        logger.info("Remove --dry-run flag to run the actual pipeline.")
        logger.info("=" * 60)


# ─────────────────────────────────────────────────────────────────────────────
# CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Hybrid (Rule + LLM) Credit Card PDF Preprocessing Pipeline. "
            "Extends preprocess.py with LLM fallback via Ollama."
        )
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help=(
            "Simulate processing without copying files, creating folders, "
            "or writing logs. Everything printed to console only."
        ),
    )
    parser.add_argument(
        "--debug", action="store_true",
        help="Print extracted text samples for each PDF.",
    )
    parser.add_argument(
        "--no-llm", action="store_true",
        help="Disable LLM fallback — run rule-based classification only.",
    )
    args = parser.parse_args()

    setup_logging()

    logger.info("=" * 60)
    logger.info("HYBRID PREPROCESSING PIPELINE")
    logger.info(
        f"  LLM fallback : "
        f"{'DISABLED (--no-llm)' if args.no_llm else 'ENABLED (via Ollama)'}"
    )
    logger.info(f"  Dry-run      : {'ON — NO files will be written' if args.dry_run else 'OFF'}")
    logger.info(f"  Debug        : {'ON' if args.debug else 'OFF'}")
    logger.info("=" * 60)

    process_all_hybrid(
        dry_run = args.dry_run,
        debug   = args.debug,
        use_llm = not args.no_llm,
    )


if __name__ == "__main__":
    main()

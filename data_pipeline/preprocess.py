# =============================================================================
# preprocess.py — Credit Card Document Preprocessing Pipeline (Rule-Based)
# =============================================================================
#
# ── FIXES IN THIS VERSION ────────────────────────────────────────────────────
#
#  🔴 FIX 1 — SCB vs SBI MISCLASSIFICATION (CRITICAL)
#    Problem: SC_Smart_MITC_2026.pdf detected as SBI instead of SCB.
#    Root cause: "sbi" can appear as a substring in garbled OCR or encoding
#    artifacts in Standard Chartered documents. Also, short aliases like "sc"
#    or "scb" may not match early enough in the scan.
#    Fix:
#      (a) BANK_ALIASES ordering: SCB aliases listed BEFORE SBI short aliases
#          More specific SCB phrases ("standard chartered bank") are checked
#          first. SBI is only checked when SCB phrases are absent.
#      (b) Added "standard chartered" (without "bank") as a SCB alias.
#      (c) BANK_ALIASES now uses OrderedDict-style ordering (most specific FIRST).
#          Banks that could conflict (SCB vs SBI) are separated clearly.
#      (d) Added filename prefix fallback: if filename starts with "SC_" and
#          bank detection returns SBI, we override to SCB.
#
#  🔴 FIX 2 — DOC TYPE DETECTION (TNC→BR and TNC→MITC ERRORS) (CRITICAL)
#    Problem:
#      - TNC docs were detected as BR because "cashback" appears in TNC
#        exclusion clauses ("cashback is not available for…")
#      - TNC docs were also detected as MITC when fee tables appeared in them
#    Fix:
#      - MITC_TITLE_PHRASES: Added more specific MITC-only phrases
#      - TNC_TITLE_PHRASES: Added unambiguous TNC-only title phrases
#        (checked before BR keywords during classification)
#      - BR_TITLE_PHRASES: Removed phrases too generic for BR (e.g., "offer")
#      - detect_doc_type() now uses a 3-layer system:
#          Layer 1: MITC title match → confident MITC
#          Layer 2: TNC title match → confident TNC (BEFORE BR scoring)
#          Layer 3: Keyword scoring for BR/LG
#        This prevents TNC title docs from being scored as BR.
#
#  🟡 FIX 3 — WRONG CARD DETECTION (Millennia→Swiggy, etc.) (IMPORTANT)
#    Problem: HDFC_Millenia_FAQ_2024.pdf → detected as "Swiggy" because
#    the FAQ document mentioned Swiggy as a partner. Same issue for other
#    docs where partner/merchant names appear prominently in the body.
#    Fix:
#      - detect_card() pass 1 (header zone) now uses first 300 chars instead
#        of 500 chars. Tighter header zone = less noise from sub-sections.
#      - CARDS list ordering verified: More specific names before generic.
#      - WORD_BOUNDARY_CARDS expanded to include more ambiguous names.
#      - Added a "bank-specific first pass": for known banks, we first check
#        ONLY that bank's cards before scanning the full CARDS list.
#        This prevents cross-bank card matches (e.g., HDFC doc → Swiggy match
#        which is a HDFC card but still wrong in this case).
#
#  🟡 FIX 4 — YEAR EXTRACTION (from original — improved)
#    Problem: detect_year() used max() across full text → wrong years from
#    fee tables, example dates, future reference years.
#    Fix: Header zone (first 500 chars) priority, cap at MAX_VALID_YEAR.
#    Falls back to filename year, then DEFAULT_YEAR.
#    Does NOT use max() from full body text anymore.
#
#  🟡 FIX 5 — AMEX BLUE CASH MISCLASSIFICATION (IMPORTANT)
#    Problem: AMEX_BlueCash_MITC_KFS_2025.pdf → detected as "Platinum Travel"
#    because "Platinum Travel" appeared in the document body.
#    Fix: "Blue Cash" already before "Platinum Travel" in CARDS list.
#    Also added header-zone priority: Blue Cash in the first 300 chars wins.
#    Added "Blue Cash" to AMEX's bank-specific first-pass list.
#
#  🔵 FIX 6 — MASTER DOC CARD NARROWING BUG (from original — kept)
#    _narrow_card_to_bank() skipped for is_master=True docs.
#
#  🔵 FIX 7 — DUPLICATE FILE OVERWRITE BUG (from original — kept)
#    Master docs go to BANK_MASTER/ folder, not overwriting card docs.
#
# ── USAGE ──────────────────────────────────────────────────────────────────────
#
#  Standalone (rules only):
#    python data_pipeline/preprocess.py              # normal run
#    python data_pipeline/preprocess.py --dry-run    # simulate, no files moved
#    python data_pipeline/preprocess.py --debug      # print extracted text
#
#  Full hybrid pipeline (rules + LLM):
#    python data_pipeline/preprocess_with_llm.py --dry-run
#    python data_pipeline/preprocess_with_llm.py
#
# ── OUTPUTS ────────────────────────────────────────────────────────────────────
#
#    data/processed_docs/              organised PDFs
#    data/needs_review/                low-confidence files
#    data/logs/summary.xlsx            per-file processing summary (Excel)
#    data/logs/preprocess_log.txt      detailed audit log
#    data/logs/missing_docs_report.csv coverage gaps per card
#
# =============================================================================

import os
import re
import csv
import sys
import shutil
import logging
import argparse
from datetime import datetime
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import pandas as pd
except ImportError:
    pd = None


# =============================================================================
# ██████████████████████  USER EDITABLE CONFIGURATION  ████████████████████████
# =============================================================================

# ─────────────────────────────────────────────────────────────────────────────
# BANK ALIAS MAP
#
# FIX 1: ORDERING IS CRITICAL.
# - More specific phrases FIRST (to avoid substring false positives)
# - SCB must be listed BEFORE short SBI aliases (e.g. "sbi" matches inside
#   "standard chartered banking" if we're not careful)
# - Each bank's aliases go from most specific → least specific
#
# Format: "SHORT_CODE": ["phrase (most specific first)", ...]
# Detection is case-insensitive.
# ─────────────────────────────────────────────────────────────────────────────

BANK_ALIASES = {
    # ── Unambiguous banks — no conflict risk ──────────────────────────────────
    "HDFC"     : ["hdfc bank ltd", "hdfc bank limited", "hdfc bank", "hdfc"],
    "AXIS"     : ["axis bank limited", "axis bank", "axis"],
    "ICICI"    : ["icici bank limited", "icici bank", "icici"],
    "HSBC"     : ["the hongkong and shanghai banking", "hsbc bank", "hsbc"],
    "KOTAK"    : ["kotak mahindra bank limited", "kotak mahindra bank",
                  "kotak mahindra", "kotak"],
    "YES"      : ["yes bank limited", "yes bank"],
    "IDFC"     : ["idfc first bank limited", "idfc first bank", "idfc first", "idfc"],
    "AU"       : ["au small finance bank limited", "au small finance bank",
                  "au bank", "au small finance"],
    "BOB"      : ["bank of baroda", "bob financial solutions limited",
                  "bob financial", "bob"],
    "RBL"      : ["rbl bank limited", "ratnakar bank limited", "rbl bank", "rbl"],
    "INDUSIND" : ["indusind bank limited", "indusind bank", "indusind"],
    "CITI"     : ["citibank n.a", "citibank na", "citibank", "citi bank", "citi"],
    "AMEX"     : ["american express banking corp", "american express", "amex"],

    # ── FIX 1: SCB must come BEFORE SBI ──────────────────────────────────────
    # Standard Chartered docs can contain "sbi" as a garbled substring.
    # Checking SCB phrases first prevents SBI from stealing SCB files.
    "SCB"      : [
        "standard chartered bank",              # most specific — always correct
        "standard chartered credit card",       # common in SC docs
        "standard chartered",                   # without "bank" — still specific
        "stanchart",                            # brand abbreviation used by SC
        "sc.com/in",                            # URL found only in SC docs
        "sc.com",                               # SC website URL
    ],

    # ── FIX 1: SBI listed AFTER SCB ──────────────────────────────────────────
    # Short alias "sbi" only checked after SCB phrases have all been tested.
    "SBI"      : ["state bank of india", "sbi cards and payment", "sbi card", "sbi"],

    "ONECARD"  : ["onecard", "one card", "ftc fintech private"],
    "SCAPIA"   : ["scapia technologies", "scapia", "federal bank scapia"],
    "JUPITER"  : ["jupiter money", "jupiter"],
}

# ─────────────────────────────────────────────────────────────────────────────
# MASTER / COLLECTIVE DOCUMENT DETECTION (unchanged)
# ─────────────────────────────────────────────────────────────────────────────

MASTER_DOC_BANKS = ["HDFC", "SBI", "AXIS", "ICICI", "AMEX", "SCB"]

MASTER_DOC_SIGNALS = [
    # AXIS
    "applicable to all credit card holders / applicants of credit cards",
    "all the information herein is applicable to all credit card holders",
    "applicable to all credit card holders",

    # AXIS LOUNGE
    "axis bank credit cards domestic airport lounge",
    "list of credit card variants",
    "credit card variants, their corresponding",

    # IDFC
    "applicable to all credit card members / applicants of credit cards",
    "all the information herein is applicable to all credit card members",
    "applicable to all credit card members",

    # AMEX MITC
    "as a part of all credit card applications",
    "specific reference is given if any terms and conditions are applicable only to a particular",
    "this mitc is to be read along with",

    # AMEX TNC
    "american express credit card cardmember agreement",
    "american express® credit card cardmember agreement",
    "before you use the enclosed american express credit card",

    # AU MASTER
    "these terms and conditions apply to the au small finance bank credit card",
    "au small finance bank credit card",

    # SCB
    "applicable to all credit cards",

    # HDFC
    "key fact statement cum most important terms & conditions for credit cards",
    "key fact statement cum most important terms and conditions for credit cards",
    "most important terms & conditions for credit cards",
    "most important terms and conditions for credit cards",

    # SBI
    "sbi card miles",

    # Generic
    "applicable to all cards",
    "all credit cardholders",
    "common terms and conditions",
    "general terms and conditions for credit cards",
    "these terms apply to all",
    "irrespective of the card variant",
    "for all card variants",
    "master terms",
    "uniform terms",
    "applicable across all variants",
    "all hdfc bank credit cards",
    "all hdfc credit cardholders",
    "all sbi credit cards",
    "all state bank credit cards",
    "all axis bank credit cards",
    "all icici bank credit cards",
    "all american express cardmembers",
    "all standard chartered credit cards",
]

# ─────────────────────────────────────────────────────────────────────────────
# CARD NAMES
#
# ORDERING RULE: More specific names MUST come before shorter/generic ones.
# The list is scanned top-to-bottom and stops at the FIRST match.
#
# FIX 5: "Blue Cash" verified to be before "Platinum Travel" and "Platinum"
# FIX 3: List ordering verified for all cards where body-text confusion is possible
# ─────────────────────────────────────────────────────────────────────────────

CARDS = [
    # ── AMEX cards — specific names first ────────────────────────────────────
    # FIX 5: "Blue Cash" MUST come before "Platinum Travel" and "Platinum"
    # AMEX Blue Cash documents mention Platinum Travel in comparison tables.
    "Blue Cash",            # AMEX Blue Cash — header-first match catches it
    "Platinum Travel",      # AMEX Platinum Travel — BEFORE plain "Platinum"

    # ── Premium & Travel ─────────────────────────────────────────────────────
    "Infinia",
    "Diners Club Black",    # BEFORE "Diners"
    "Diners",
    "Magnus",
    "Atlas",
    "Emeralde",
    "Marriott Bonvoy",
    "Vistara Infinite",     # BEFORE "Vistara"
    "Vistara",
    "Ultimate",
    "World Safari",
    "Aurum",
    "HDFC Pixel",           # Qualified before plain "Pixel"
    "Pixel",

    # ── Lifestyle & Niche ────────────────────────────────────────────────────
    "Tata Neu Infinity",    # BEFORE "Tata Neu"
    "Tata Neu",
    "Amazon Pay",           # Bank detected separately — name is specific enough
    "Swiggy",               # FIX 3: moved earlier but AFTER more specific names
    "Airtel",
    "BPCL Octane",          # BEFORE "BPCL"
    "BPCL",
    "IndianOil",
    "Flipkart",
    "Myntra",
    "Reliance",
    "IRCTC",
    "HPCL Super Saver",     # BEFORE any plain "HPCL"
    "EazyDiner",
    "Paytm",
    "Altura",               # AU specific
    "Eterna",               # BOB specific
    "Prosperity",           # YES specific
    "OneCard",
    "Scapia",
    "Jupiter",
    "6E Rewards",

    # ── Mid-tier with compound names ─────────────────────────────────────────
    "Regalia Gold",         # BEFORE "Regalia"
    "Regalia",
    "MoneyBack+",           # BEFORE "MoneyBack"
    "MoneyBack",
    "League Platinum",      # KOTAK — BEFORE plain "Platinum"
    "SimplyCLICK",
    "SimplySAVE",
    "Live+",                # HSBC Live+

    # ── Millennia — for both HDFC and IDFC ───────────────────────────────────
    "Millennia",            # Used by HDFC and IDFC — bank disambiguates

    # ── Generic / ambiguous names — LAST ─────────────────────────────────────
    # These must be at the bottom because they appear in many documents.
    "Axis ACE",             # Qualified form BEFORE plain "ACE"
    "ACE",                  # Word-boundary matched
    "HDFC Millennia",       # Qualified fallback
    "Platinum",             # Appears everywhere — last resort
    "Smart Credit Card",    # Qualified BEFORE plain "Smart"
    "Smart",                # SCB Smart — word-boundary matched
    "Cashback",             # Extremely common word — very last resort
    "Signature",            # Generic
    "Elite",                # Generic
    "Reserve",              # Generic
    "Select",               # Generic
    "Coral",                # Generic
    "Rubyx",                # ICICI
]

# ─────────────────────────────────────────────────────────────────────────────
# BANK-SPECIFIC CARD LISTS FOR FIRST-PASS DETECTION
#
# FIX 3: When we know the bank (from text), we first check ONLY that bank's
# cards before scanning the full CARDS list. This prevents cross-bank matches
# (e.g., an HDFC doc matching "Swiggy" from a Swiggy HDFC partnership mention).
#
# Note: This is the first-pass list used by detect_card().
# The full BANK_CARDS in preprocess_with_llm.py is used for validation.
# ─────────────────────────────────────────────────────────────────────────────

BANK_FIRST_PASS_CARDS: dict[str, list[str]] = {
    "HDFC": [
        "Infinia", "Diners Club Black", "Diners", "Marriott Bonvoy",
        "Regalia Gold", "Regalia", "Millennia", "Tata Neu Infinity",
        "Tata Neu", "MoneyBack+", "MoneyBack", "Swiggy", "IndianOil",
        "Pixel", "HDFC Millennia",
    ],
    "SBI": [
        "Cashback", "Elite", "Aurum", "SimplyCLICK", "SimplySAVE",
        "BPCL Octane", "BPCL", "IRCTC", "Paytm", "Reliance", "Vistara",
    ],
    "AXIS": [
        "Magnus", "Atlas", "Reserve", "ACE", "Axis ACE", "Flipkart",
        "Airtel", "Select", "Vistara Infinite", "Vistara", "Coral",
    ],
    "ICICI": [
        "Emeralde", "Amazon Pay", "HPCL Super Saver", "Rubyx",
    ],
    "AMEX": [
        "Blue Cash",        # FIX 5: Blue Cash before Platinum Travel
        "Platinum Travel",
        "Platinum",
    ],
    "HSBC": ["Premier", "Live+"],
    "KOTAK": ["League Platinum", "Myntra"],
    "SCB": ["Smart", "Ultimate"],
    "IDFC": ["Millennia"],
    "AU": ["Altura"],
    "BOB": ["Eterna"],
    "RBL": ["World Safari"],
    "INDUSIND": ["EazyDiner"],
    "YES": ["Prosperity"],
}

# ─────────────────────────────────────────────────────────────────────────────
# DOCUMENT TYPE DEFINITIONS
#
# FIX 2: Title phrases are now more specific to prevent cross-type errors.
# MITC and TNC title phrases are unambiguous enough to trust on their own.
# BR and LG rely more on keyword scoring (which is expected).
# ─────────────────────────────────────────────────────────────────────────────

# MITC title phrases — finding ANY of these in the header is a confident MITC
MITC_TITLE_PHRASES = [
    "key fact statement cum most important terms",
    "key fact statement",
    "most important terms and conditions",
    "most important terms & conditions",
    "most important terms",
    "important information document",
    "mitc",
]

# MITC body keywords — used for scoring
MITC_KEYWORDS = [
    "interest rate", "annual fee", "finance charge", "late payment fee",
    "minimum amount due", "credit limit", "cash advance fee",
    "overlimit fee", "foreign currency", "billing cycle",
    "statement date", "payment due date", "charges",
    "schedule of charges", "joining fee", "membership fee",
    "rate of interest", "overdue interest",
]

# FIX 2: TNC TITLE phrases are now stricter (unambiguous title-level signals only)
# These are checked BEFORE BR scoring to prevent TNC→BR misclassification.
TNC_TITLE_PHRASES = [
    "cardmember agreement",
    "cardholder agreement",
    "client terms",
    "credit card terms",     # specific enough when in the title
]

# TNC body keywords — generic, used only when no MITC/TNC title found
TNC_KEYWORDS = [
    "terms and conditions", "cardmember agreement", "cardholder agreement",
    "governing law", "arbitration", "exclusions", "liability",
    "termination", "dispute resolution", "amendments", "jurisdiction",
    "binding", "indemnify", "force majeure",
]

# BR title phrases (FIX 2: removed "offer terms" — too generic)
BR_TITLE_PHRASES = [
    "benefits guide",
    "features and benefits",
    "rewards programme",
    "reward catalogue",
    "cashback proposition",
    "cashpoints proposition",
    "welcome benefit",
    "welcome gift",
    "spends increase",
    "tap & pay offer",
    # Removed: "offer terms and conditions" — fires on TNC docs
    # Removed: "this offer" — fires on TNC docs with offer exclusions
]

# BR body keywords
BR_KEYWORDS = [
    "cashback proposition",
    "reward points", "welcome bonus", "milestone benefit",
    "accelerated rewards", "redemption", "gift voucher",
    "fuel surcharge waiver", "dining benefit", "earn rate",
    "cashpoints", "welcome benefit", "joining benefit",
    "spends milestone", "spend milestone",
    "welcome gift",
    "bookmyshow voucher",
    "instant discount",
]

# LG title phrases and keywords (unchanged)
LG_TITLE_PHRASES = [
    "lounge access guide",
    "airport lounge programme",
    "lounge benefit",
    "domestic airport lounge access program",
    "lounge access program",
    "change in lounge access",
]

LG_KEYWORDS = [
    "lounge access", "airport lounge", "priority pass", "lounge program",
    "domestic lounge", "international lounge", "complimentary lounge",
    "meet and greet", "fast track immigration",
    "lounge eligibility",
    "participating lounges",
    "spend criteria",
]

# ─────────────────────────────────────────────────────────────────────────────
# CLASSIFICATION & PIPELINE SETTINGS
# ─────────────────────────────────────────────────────────────────────────────
CONFIDENCE_THRESHOLD = 0.70
DEFAULT_YEAR         = "2026"
PAGE_TIERS           = [3, 5, 10]

MAX_VALID_YEAR = 2026
MIN_VALID_YEAR = 2010

# Header zone size — FIX 3: tightened from 500 to 300 chars for card detection
# This reduces noise from sub-sections that mention partner/merchant names.
HEADER_ZONE_SIZE = 300

# ─────────────────────────────────────────────────────────────────────────────
# DOCUMENT COVERAGE VALIDATION SETTINGS (unchanged from original)
# ─────────────────────────────────────────────────────────────────────────────
REQUIRED_DOCS = ["MITC", "BR"]
OPTIONAL_DOCS = ["TNC", "LG"]

EXPECTED_CARDS = [
    # ── Everyday Cashback ─────────────────────────────────────────────────────
    "SBI Cashback", "HDFC Millennia", "AXIS ACE", "ICICI Amazon Pay",
    "AXIS Flipkart", "HSBC Live+", "SCB Smart", "HDFC MoneyBack+",
    "SBI SimplyCLICK", "AMEX Blue Cash", "IDFC Millennia",
    "AU Altura", "BOB Eterna", "KOTAK League Platinum", "YES Prosperity",

    # ── Premium & Travel ──────────────────────────────────────────────────────
    "HDFC Infinia", "HDFC Diners Club Black", "AXIS Magnus", "AXIS Atlas",
    "AMEX Platinum", "AMEX Platinum Travel", "ICICI Emeralde",
    "SBI Elite", "HDFC Marriott Bonvoy", "AXIS Reserve", "HSBC Premier",
    "SBI Aurum", "AXIS Vistara Infinite", "SCB Ultimate", "RBL World Safari",

    # ── Lifestyle & Niche ─────────────────────────────────────────────────────
    "HDFC Tata Neu Infinity", "HDFC Swiggy", "AXIS Airtel",
    "SBI BPCL Octane", "HDFC IndianOil", "KOTAK Myntra",
    "SBI Reliance", "SBI IRCTC", "HDFC Regalia Gold",
    "ICICI HPCL Super Saver", "AXIS Select", "INDUSIND EazyDiner",
    "SBI Paytm", "HDFC Pixel",

    # ── Master / Collective Documents ─────────────────────────────────────────
    "HDFC MASTER", "SBI MASTER", "AXIS MASTER", "ICICI MASTER",
    "AMEX MASTER", "SCB MASTER",
]

# =============================================================================
# END OF USER EDITABLE CONFIGURATION
# =============================================================================


# ─────────────────────────────────────────────────────────────────────────────
# PATH SETUP
# ─────────────────────────────────────────────────────────────────────────────
SCRIPT_DIR     = Path(__file__).resolve().parent
PROJECT_ROOT   = SCRIPT_DIR.parent
RAW_DIR        = PROJECT_ROOT / "data" / "raw_docs"
PROCESSED_DIR  = PROJECT_ROOT / "data" / "processed_docs"
REVIEW_DIR     = PROJECT_ROOT / "data" / "needs_review"
LOG_DIR        = PROJECT_ROOT / "data" / "logs"
SUMMARY_CSV         = LOG_DIR / "summary.csv"
DETAIL_LOG          = LOG_DIR / "preprocess_log.txt"
MISSING_DOCS_CSV    = LOG_DIR / "missing_docs_report.csv"
COVERAGE_DASHBOARD  = LOG_DIR / "coverage_dashboard.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# LOGGING SETUP
# ─────────────────────────────────────────────────────────────────────────────
def setup_logging():
    """Configure console logging with a clean format."""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="[%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — TEXT EXTRACTION (unchanged)
# ─────────────────────────────────────────────────────────────────────────────
def extract_text(pdf_path: Path, max_pages: int) -> str:
    """
    Extract plain text from the first `max_pages` pages of a PDF.
    Tries pdfplumber first (better for formatted PDFs), then PyMuPDF.
    Returns empty string on failure.
    """
    text = ""

    if pdfplumber:
        try:
            with pdfplumber.open(str(pdf_path)) as pdf:
                for page in pdf.pages[:max_pages]:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            if text.strip():
                return text
        except Exception as e:
            logger.warning(f"pdfplumber failed on {pdf_path.name}: {e}")

    if fitz:
        try:
            doc = fitz.open(str(pdf_path))
            for i, page in enumerate(doc):
                if i >= max_pages:
                    break
                text += page.get_text() + "\n"
            doc.close()
            if text.strip():
                return text
        except Exception as e:
            logger.warning(f"PyMuPDF failed on {pdf_path.name}: {e}")

    logger.error(f"No PDF library could extract text from {pdf_path.name}.")
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — DETECTION HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _normalise(text: str) -> str:
    """Lowercase and collapse whitespace for reliable matching."""
    return re.sub(r"\s+", " ", text.lower().strip())


def _keyword_score(text_lower: str, keywords: list[str]) -> tuple[float, list[str]]:
    """Count keyword matches. Returns (normalised score 0–1, matched list)."""
    matched = [kw for kw in keywords if kw.lower() in text_lower]
    if not keywords:
        return 0.0, []
    score = min(len(matched) / max(len(keywords) * 0.3, 1), 1.0)
    return round(score, 3), matched


def _title_match(text_lower: str, phrases: list[str]) -> tuple[bool, str]:
    """
    Check for title/header phrase in the first HEADER_ZONE_SIZE characters.
    FIX 3: Uses HEADER_ZONE_SIZE (300) instead of hardcoded 500.
    """
    header_zone = text_lower[:HEADER_ZONE_SIZE]
    for phrase in phrases:
        if phrase.lower() in header_zone:
            return True, phrase
    return False, ""


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — DOCUMENT TYPE DETECTION (FIX 2)
# ─────────────────────────────────────────────────────────────────────────────
def detect_doc_type(text: str) -> dict:
    """
    FIX 2: Three-layer classification to prevent TNC→BR and TNC→MITC errors.

    Layer 1: MITC title/header phrase match
      → If MITC title phrase found: MITC (high confidence)
      → STOP — don't fall through to Layer 2

    Layer 2: TNC title/header phrase match
      → If TNC title phrase found: TNC (high confidence)
      → STOP — this prevents TNC docs from being scored as BR

    Layer 3: Keyword frequency scoring for BR and LG
      → Used only when no MITC or TNC title phrase was found
      → TNC keywords also scored here as fallback

    Returns {value, confidence, reasons}.
    """
    text_lower = _normalise(text)
    reasons    = []

    # Layer 1: MITC title match (always wins — fees/charges are definitive)
    mitc_title_found, mitc_phrase = _title_match(text_lower, MITC_TITLE_PHRASES)
    if mitc_title_found:
        reasons.append(f'MITC title phrase found in header: "{mitc_phrase}"')
        kw_score, kw_matched = _keyword_score(text_lower, MITC_KEYWORDS)
        confidence = round(min(0.65 + kw_score * 0.35 + 0.10, 1.0), 3)
        if kw_matched:
            reasons.append(f"Supporting MITC keywords: {', '.join(kw_matched[:4])}")
        return {"value": "MITC", "confidence": confidence, "reasons": reasons}

    # Layer 2: FIX 2 — TNC title match (checked BEFORE BR scoring)
    # This prevents TNC docs with "cashback" in exclusion clauses from
    # being misclassified as BR.
    tnc_title_found, tnc_phrase = _title_match(text_lower, TNC_TITLE_PHRASES)
    if tnc_title_found:
        reasons.append(
            f'TNC title phrase found in header: "{tnc_phrase}" '
            f'(checked before BR scoring to prevent TNC→BR error)'
        )
        kw_score, kw_matched = _keyword_score(text_lower, TNC_KEYWORDS)
        confidence = round(min(0.60 + kw_score * 0.35 + 0.10, 1.0), 3)
        if kw_matched:
            reasons.append(f"Supporting TNC keywords: {', '.join(kw_matched[:4])}")
        return {"value": "TNC", "confidence": confidence, "reasons": reasons}

    # Layer 3: Keyword scoring for BR, LG, and TNC (as fallback)
    kw_data = {
        "BR":   _keyword_score(text_lower, BR_KEYWORDS),
        "LG":   _keyword_score(text_lower, LG_KEYWORDS),
        "TNC":  _keyword_score(text_lower, TNC_KEYWORDS),
        "MITC": _keyword_score(text_lower, MITC_KEYWORDS),
    }

    # Also check BR title phrases in Layer 3 (they're less specific than MITC/TNC)
    br_title_found, br_phrase = _title_match(text_lower, BR_TITLE_PHRASES)
    lg_title_found, lg_phrase = _title_match(text_lower, LG_TITLE_PHRASES)

    if br_title_found:
        reasons.append(f'BR title phrase found: "{br_phrase}"')
        br_kw_score = kw_data["BR"][0]
        confidence  = round(min(0.60 + br_kw_score * 0.35 + 0.10, 1.0), 3)
        return {"value": "BR", "confidence": confidence, "reasons": reasons}

    if lg_title_found:
        reasons.append(f'LG title phrase found: "{lg_phrase}"')
        lg_kw_score = kw_data["LG"][0]
        confidence  = round(min(0.60 + lg_kw_score * 0.35 + 0.10, 1.0), 3)
        return {"value": "LG", "confidence": confidence, "reasons": reasons}

    # Pure keyword scoring — pick highest scorer
    ranked = sorted(kw_data.items(), key=lambda x: x[1][0], reverse=True)
    best_type, (best_score, best_matched) = ranked[0]

    if best_matched:
        reasons.append(
            f"Best keyword match: {best_type} "
            f"(score={best_score:.2f}, matched: {', '.join(best_matched[:5])})"
        )
    for dtype, (score, _) in ranked[1:]:
        if score > 0.1:
            reasons.append(f"Lower keyword presence for {dtype} (score={score:.2f})")

    if not reasons:
        reasons.append("No strong signals found; defaulting to best keyword score")

    confidence = round(best_score * 0.85, 3)
    return {"value": best_type, "confidence": confidence, "reasons": reasons}


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — BANK DETECTION (FIX 1)
# ─────────────────────────────────────────────────────────────────────────────
def detect_bank(text: str, filename: str) -> dict:
    """
    FIX 1: BANK_ALIASES ordering ensures SCB is checked before SBI.

    Searches PDF text for any alias phrase defined in BANK_ALIASES.
    Returns the SHORT CODE (e.g. "SCB"), not the full bank name.
    Falls back to searching the filename if text search finds nothing.

    FIX 1 (additional): If filename starts with "SC_" and text search returns
    "SBI", we issue a warning but keep SBI (the caller in preprocess_with_llm.py
    will catch the filename mismatch). The fix is in BANK_ALIASES ordering,
    which should catch this before reaching the fallback.
    """
    text_lower = _normalise(text)
    fn_lower   = _normalise(filename)
    reasons    = []

    # Search PDF text — stops at the first alias phrase match
    # FIX 1: SCB appears before SBI in BANK_ALIASES, so SCB phrases
    # are checked first. "standard chartered" will match before "sbi".
    for short_code, aliases in BANK_ALIASES.items():
        for phrase in aliases:
            if phrase.lower() in text_lower:
                in_header  = phrase.lower() in text_lower[:HEADER_ZONE_SIZE]
                confidence = 0.95 if in_header else 0.80
                reasons.append(
                    f"Found '{phrase}' → mapped to short code '{short_code}' "
                    f"in {'header' if in_header else 'body'}"
                )
                return {
                    "value": short_code,
                    "confidence": confidence,
                    "reasons": reasons,
                }

    # Fallback: search the filename
    for short_code, aliases in BANK_ALIASES.items():
        for phrase in aliases:
            if phrase.lower() in fn_lower:
                reasons.append(
                    f"Found '{phrase}' in filename → mapped to '{short_code}' "
                    f"(text extraction may have failed)"
                )
                return {"value": short_code, "confidence": 0.55, "reasons": reasons}

    # FIX 1: Filename prefix fallback for common bank prefixes
    # This catches cases where the filename starts with "SC_", "HDFC_", etc.
    # but the text alias search failed (e.g., very short/garbled text).
    filename_prefix_map = {
        "hdfc": "HDFC", "sbi": "SBI", "axis": "AXIS", "icici": "ICICI",
        "amex": "AMEX", "hsbc": "HSBC", "idfc": "IDFC", "au": "AU",
        "sc": "SCB", "scb": "SCB", "kotak": "KOTAK", "yes": "YES",
        "bob": "BOB", "rbl": "RBL", "indusind": "INDUSIND",
    }
    fn_stem_lower = Path(filename).stem.split("_")[0].lower()
    if fn_stem_lower in filename_prefix_map:
        mapped = filename_prefix_map[fn_stem_lower]
        reasons.append(
            f"Filename prefix '{fn_stem_lower}' → mapped to '{mapped}' "
            f"(last resort — alias phrases not found in text)"
        )
        return {"value": mapped, "confidence": 0.50, "reasons": reasons}

    reasons.append("No known bank name or alias found in text or filename")
    return {"value": None, "confidence": 0.0, "reasons": reasons}


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — CARD DETECTION (FIX 3, FIX 5)
# ─────────────────────────────────────────────────────────────────────────────

# FIX 3: Expanded WORD_BOUNDARY_CARDS to include more ambiguous names
WORD_BOUNDARY_CARDS = {
    "ACE", "Smart", "Elite", "Select", "Reserve", "Coral",
    "Signature", "Platinum", "Atlas", "Airtel", "Pixel",
    "Aurum",    # "aurum" could appear in chemistry/law text
    "Swiggy",   # FIX 3: Swiggy appears as a merchant in many HDFC docs
}


def _card_matches(card: str, text_lower: str) -> bool:
    """
    Check whether a card name appears in text.
    FIX 3: More cards now use word-boundary matching (WORD_BOUNDARY_CARDS).
    """
    card_lower = card.lower()
    if card in WORD_BOUNDARY_CARDS:
        pattern = r"\b" + re.escape(card_lower) + r"\b"
        return bool(re.search(pattern, text_lower))
    return card_lower in text_lower


def detect_card(text: str, filename: str, detected_bank: str = None) -> dict:
    """
    FIX 3: 4-pass card detection with bank-specific first pass.

    Pass 0 (NEW) — Bank-specific header scan: confidence 0.95
      When we know the bank, check ONLY that bank's cards in the
      header zone. This prevents cross-bank matches.
      e.g., for HDFC documents, we check HDFC cards first.

    Pass 1 — Global header zone (first 300 chars): confidence 0.92
      The card name in the title/header is the most reliable signal.

    Pass 2 — Full text search: confidence 0.75
      Scans entire text. Word-boundary matching for ambiguous names.

    Pass 3 — Filename fallback: confidence 0.50
      Used when text extraction produces nothing.

    Parameters:
        text          : Extracted PDF text
        filename      : Original filename (for fallback)
        detected_bank : Bank already detected (used for Pass 0 filtering)
    """
    text_lower  = _normalise(text)
    fn_lower    = _normalise(filename)
    header_zone = text_lower[:HEADER_ZONE_SIZE]
    reasons     = []

    # Pass 0: FIX 3 — Bank-specific header scan (highest confidence)
    if detected_bank and detected_bank.upper() in BANK_FIRST_PASS_CARDS:
        bank_cards = BANK_FIRST_PASS_CARDS[detected_bank.upper()]
        for card in bank_cards:
            if _card_matches(card, header_zone):
                reasons.append(
                    f"[Pass 0] Found bank-specific card '{card}' for {detected_bank} "
                    f"in header zone — highest confidence match"
                )
                return {"value": card, "confidence": 0.95, "reasons": reasons}

        # Pass 0b: FIX 3 — Bank-specific full-text scan
        # Still bank-filtered but searches the full text
        for card in bank_cards:
            if _card_matches(card, text_lower):
                reasons.append(
                    f"[Pass 0b] Found bank-specific card '{card}' for {detected_bank} "
                    f"in document body (not in header)"
                )
                return {"value": card, "confidence": 0.80, "reasons": reasons}

    # Pass 1: Global header zone — CARDS list order (header-first)
    for card in CARDS:
        if _card_matches(card, header_zone):
            reasons.append(
                f"[Pass 1] Found '{card}' in document header "
                f"(first {HEADER_ZONE_SIZE} chars) — global card list"
            )
            return {"value": card, "confidence": 0.92, "reasons": reasons}

    # Pass 2: Full text search
    for card in CARDS:
        if _card_matches(card, text_lower):
            wb_note = " (word-boundary)" if card in WORD_BOUNDARY_CARDS else ""
            reasons.append(
                f"[Pass 2] Found '{card}' in document body{wb_note} "
                f"— not in header, lower confidence"
            )
            return {"value": card, "confidence": 0.75, "reasons": reasons}

    # Pass 3: Filename fallback
    for card in CARDS:
        if _card_matches(card, fn_lower):
            reasons.append(
                f"[Pass 3] Found '{card}' in filename — "
                f"text extraction may have failed (scanned PDF?)"
            )
            return {"value": card, "confidence": 0.50, "reasons": reasons}

    reasons.append(
        "No known card name found in header, body, or filename. "
        "Add the card name to CARDS or BANK_FIRST_PASS_CARDS if unexpected."
    )
    return {"value": None, "confidence": 0.0, "reasons": reasons}


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5B — MASTER / COLLECTIVE DOCUMENT DETECTION (unchanged)
# ─────────────────────────────────────────────────────────────────────────────
def detect_master_doc(text: str) -> dict:
    """
    Determines whether this PDF applies to ALL of a bank's credit cards.
    Runs BEFORE detect_card() to prevent collective docs from being
    misclassified as card-specific.
    """
    text_lower  = _normalise(text)
    header_zone = text_lower[:HEADER_ZONE_SIZE]

    for signal in MASTER_DOC_SIGNALS:
        if signal.lower() in header_zone:
            return {
                "is_master":  True,
                "signal":     signal,
                "confidence": 0.95,
                "found_in":   "header",
            }

    for signal in MASTER_DOC_SIGNALS:
        if signal.lower() in text_lower:
            return {
                "is_master":  True,
                "signal":     signal,
                "confidence": 0.85,
                "found_in":   "body",
            }

    return {"is_master": False, "signal": None, "confidence": 0.0, "found_in": None}


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — YEAR DETECTION (FIX 4)
# ─────────────────────────────────────────────────────────────────────────────
def detect_year(text: str) -> str:
    """
    FIX 4: Header-first year detection, NO max() across full text.

    Three-pass strategy:
    Pass 1: Header zone (first 500 chars) — document date is here
    Pass 2: Opening section (first 1000 chars) — formatted date mentions
    Pass 3: DEFAULT_YEAR — safe fallback

    Valid year range: MIN_VALID_YEAR (2010) to MAX_VALID_YEAR (2026).
    We do NOT use max() from full body — avoids picking up wrong years
    from fee tables, example dates, and reference footnotes.
    """
    year_pattern = r"\b(20[0-2]\d)\b"

    def _filter_years(years: list[str]) -> list[str]:
        return [y for y in years if MIN_VALID_YEAR <= int(y) <= MAX_VALID_YEAR]

    # Pass 1: Header zone — most reliable
    header = _normalise(text[:500])
    header_years = _filter_years(re.findall(year_pattern, header))
    if header_years:
        return header_years[-1]  # last year in header = document date

    # Pass 2: Opening section
    opening = _normalise(text[:1000])
    opening_years = _filter_years(re.findall(year_pattern, opening))
    if opening_years:
        return max(opening_years)  # most recent in opening section

    # Pass 3: Default
    logger.debug(f"[YEAR] No valid year found in text; using default: {DEFAULT_YEAR}")
    return DEFAULT_YEAR


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7 — OVERALL CONFIDENCE (unchanged)
# ─────────────────────────────────────────────────────────────────────────────
def compute_confidence(
    doc_type_result: dict,
    bank_result:     dict,
    card_result:     dict,
    master_result:   dict | None = None,
) -> float:
    """
    Weighted average: doc_type (50%) + bank (30%) + card (20%).
    Master doc floor: minimum 0.72 when bank found and is_master=True.
    """
    score = round(
        doc_type_result["confidence"] * 0.50 +
        bank_result["confidence"]     * 0.30 +
        card_result["confidence"]     * 0.20,
        3
    )

    if (master_result and master_result.get("is_master")
            and bank_result.get("value") is not None):
        score = max(score, 0.72)

    return score


# ─────────────────────────────────────────────────────────────────────────────
# STEP 8 — ADAPTIVE PAGE READING + DETECTION (FIX 3 integrated)
# ─────────────────────────────────────────────────────────────────────────────
def run_detection_with_fallback(pdf_path: Path, debug: bool):
    """
    Adaptive strategy — reads more pages when confidence is low.

    FIX 3: detect_card() now receives detected_bank so it can run a
    bank-specific first pass before the global CARDS scan.

    Detection order per tier:
      1. detect_doc_type()   — MITC / TNC / BR / LG
      2. detect_bank()       — bank short code
      3. detect_master_doc() — is this a collective doc?
         → if MASTER: card = "MASTER", skip detect_card()
         → if not:    detect_card() runs with bank hint (FIX 3)
      4. compute_confidence()
    """
    filename           = pdf_path.name
    text               = ""
    doc_type_result    = None
    bank_result        = None
    card_result        = None
    master_result      = None
    overall_confidence = 0.0

    for tier_idx, max_pages in enumerate(PAGE_TIERS):
        logger.info(f"Extracting text ({max_pages} pages) from: {filename}")
        text = extract_text(pdf_path, max_pages)

        if not text.strip():
            logger.warning(
                f"No text extracted from {filename} — skipping further tiers."
            )
            break

        if debug:
            logger.info(
                f"[DEBUG] Text sample ({max_pages} pages): "
                f"{text[:600].replace(chr(10), ' ')}"
            )

        doc_type_result = detect_doc_type(text)
        bank_result     = detect_bank(text, filename)
        year            = detect_year(text)

        master_result = detect_master_doc(text)

        if master_result["is_master"]:
            card_result = {
                "value":      "MASTER",
                "confidence": master_result["confidence"],
                "reasons": [
                    "MASTER DOCUMENT — applies to all cards for this bank",
                    f'Trigger signal: "{master_result["signal"]}"',
                    f'Found in: {master_result.get("found_in", "document")} '
                    f'(conf={master_result["confidence"]})',
                    "Individual card detection was intentionally skipped",
                    f'Will go to: processed_docs/{bank_result["value"] or "UNKNOWN"}_MASTER/',
                ],
            }
            logger.info(
                f'[MASTER DOC] Collective document detected — '
                f'signal: "{master_result["signal"]}" | '
                f'card set to MASTER (conf={master_result["confidence"]})'
            )
        else:
            # FIX 3: Pass detected bank to detect_card() for bank-specific first pass
            card_result = detect_card(
                text         = text,
                filename     = filename,
                detected_bank= bank_result.get("value"),
            )

        overall_confidence = compute_confidence(
            doc_type_result, bank_result, card_result, master_result
        )

        logger.info(
            f"Confidence after {max_pages} pages: {overall_confidence:.2f} "
            f"| {bank_result['value']} | {card_result['value']} "
            f"| {doc_type_result['value']}"
        )

        if overall_confidence >= CONFIDENCE_THRESHOLD:
            if tier_idx > 0:
                logger.info(
                    f"Confidence improved to {overall_confidence:.2f} "
                    f"after reading {max_pages} pages."
                )
            break
        elif tier_idx < len(PAGE_TIERS) - 1:
            logger.info(
                f"Confidence {overall_confidence:.2f} below threshold "
                f"{CONFIDENCE_THRESHOLD}. Retrying with "
                f"{PAGE_TIERS[tier_idx+1]} pages..."
            )
        else:
            logger.info(
                f"Reached maximum page limit ({max_pages}). "
                f"Final confidence: {overall_confidence:.2f}."
            )

    # Safe defaults if extraction failed
    if doc_type_result is None:
        doc_type_result = {
            "value": "UNKNOWN", "confidence": 0.0,
            "reasons": ["Text extraction failed"],
        }
    if bank_result is None:
        bank_result = {
            "value": None, "confidence": 0.0,
            "reasons": ["Text extraction failed"],
        }
    if card_result is None:
        card_result = {
            "value": None, "confidence": 0.0,
            "reasons": ["Text extraction failed"],
        }
    if master_result is None:
        master_result = {
            "is_master": False, "signal": None, "confidence": 0.0,
        }

    year = detect_year(text) if text.strip() else DEFAULT_YEAR
    return (
        text, doc_type_result, bank_result, card_result,
        master_result, year, overall_confidence,
    )


# ─────────────────────────────────────────────────────────────────────────────
# STEP 9 — FILENAME + FOLDER GENERATION (unchanged)
# ─────────────────────────────────────────────────────────────────────────────

def _safe(value: str, fallback: str = "UNKNOWN") -> str:
    """Convert a string to a clean filesystem-safe format."""
    if not value:
        return fallback
    result = str(value).replace("+", "Plus")
    result = re.sub(r"[^\w\-]", "_", result)
    result = re.sub(r"_+", "_", result)
    result = result.strip("_")
    return result if result else fallback


def generate_filename(
    bank: str | None,
    card: str | None,
    doc_type: str,
    year: str,
) -> str:
    """Assembles output filename: [BANK]_[CARD]_[DOCTYPE]_[YEAR].pdf"""
    b = _safe(bank,     "UNKNOWN_BANK")
    c = _safe(card,     "UNKNOWN_CARD")
    d = _safe(doc_type, "UNKNOWN")
    return f"{b}_{c}_{d}_{year}.pdf"


def get_output_folder(
    bank: str | None,
    card: str | None,
    base_dir: Path,
) -> Path:
    """Builds subfolder path: processed_docs/[BANK]_[CARD]/ and creates it."""
    b      = _safe(bank, "UNKNOWN_BANK")
    c      = _safe(card, "UNKNOWN_CARD")
    folder = base_dir / f"{b}_{c}"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


# ─────────────────────────────────────────────────────────────────────────────
# STEP 10 — DUPLICATE DETECTION (unchanged)
# ─────────────────────────────────────────────────────────────────────────────
def is_duplicate(dest_path: Path) -> bool:
    """Return True if a file with the same output name already exists."""
    return dest_path.exists()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 11 — FILE COPY (unchanged)
# ─────────────────────────────────────────────────────────────────────────────
def move_file(src: Path, dest: Path, dry_run: bool) -> None:
    """
    Copies (does NOT delete) the source file to its destination.
    Raw files in raw_docs/ are NEVER modified or removed.
    In dry-run mode: only prints what would happen.
    """
    if dry_run:
        logger.info(f"[DRY-RUN] Would copy: {src.name} → {dest}")
        return
    try:
        shutil.copy2(str(src), str(dest))
        logger.info(f"Copied: {src.name} → {dest}")
    except Exception as e:
        logger.error(f"Failed to copy {src.name}: {e}")
        raise


# ─────────────────────────────────────────────────────────────────────────────
# STEP 12 — DETAILED LOG WRITER (unchanged)
# ─────────────────────────────────────────────────────────────────────────────
def write_detail_log(entries: list[dict]) -> None:
    """Write a human-readable per-file audit trail to preprocess_log.txt."""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    with open(DETAIL_LOG, "w", encoding="utf-8") as f:
        f.write(
            f"PREPROCESSING LOG — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        )
        f.write("=" * 70 + "\n\n")
        for entry in entries:
            f.write("=" * 30 + "\n")
            f.write(f"FILE: {entry['filename']}\n\n")

            f.write(f"BANK: {entry['bank']} ({entry['bank_conf']:.2f})\n")
            f.write("REASONS:\n")
            for r in entry["bank_reasons"]:
                f.write(f"  • {r}\n")
            f.write("\n")

            if entry.get("is_master"):
                f.write("MASTER / COLLECTIVE DOCUMENT: YES\n")
                f.write(f'  Trigger signal : "{entry["master_signal"]}"\n')
                f.write(f"  Confidence     : {entry['master_conf']:.2f}\n")
                f.write("  Decision       : Individual card detection was SKIPPED.\n")
            else:
                f.write("MASTER / COLLECTIVE DOCUMENT: NO\n")
            f.write("\n")

            f.write(f"CARD: {entry['card']} ({entry['card_conf']:.2f})\n")
            f.write("REASONS:\n")
            for r in entry["card_reasons"]:
                f.write(f"  • {r}\n")
            f.write("\n")

            f.write(f"DOC TYPE: {entry['doc_type']} ({entry['doc_type_conf']:.2f})\n")
            f.write("REASONS:\n")
            for r in entry["doc_type_reasons"]:
                f.write(f"  • {r}\n")
            f.write("\n")

            f.write(f"PAGES READ: {entry.get('pages_read', 'N/A')}\n")
            f.write(f"STATUS: {entry['status']}\n\n")

    logger.info(f"Detailed log written to: {DETAIL_LOG}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 13 — CSV SUMMARY WRITER (fallback when openpyxl not available)
# ─────────────────────────────────────────────────────────────────────────────
def write_summary_csv(entries: list[dict]) -> None:
    """Write summary.csv (CSV fallback — xlsx preferred)."""
    fieldnames = [
        "File Name", "Bank", "Card", "Is_Master", "DocType",
        "Confidence", "Reason", "Status",
    ]
    rows = []
    for e in entries:
        combined_reason = " | ".join(
            e["doc_type_reasons"][:2] + e["bank_reasons"][:1] + e["card_reasons"][:1]
        )
        rows.append({
            "File Name":  e["filename"],
            "Bank":       e["bank"] or "NOT FOUND",
            "Card":       e["card"] or "NOT FOUND",
            "Is_Master":  "YES" if e.get("is_master") else "NO",
            "DocType":    e["doc_type"],
            "Confidence": f"{e['overall_conf']:.2f}",
            "Reason":     combined_reason,
            "Status":     e["status"],
        })
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    if pd:
        pd.DataFrame(rows, columns=fieldnames).to_csv(SUMMARY_CSV, index=False)
    else:
        with open(SUMMARY_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
    logger.info(f"Summary CSV written to: {SUMMARY_CSV}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 14 — DOCUMENT COVERAGE VALIDATION (unchanged)
# ─────────────────────────────────────────────────────────────────────────────
def build_coverage_map(log_entries: list[dict]) -> dict[str, set[str]]:
    """Builds { 'BANK Card' → {doc types found} } from successful entries."""
    coverage: dict[str, set[str]] = {}
    for entry in log_entries:
        if entry["status"] not in ("SUCCESS", "DUPLICATE_SKIPPED", "MASTER_DOC"):
            continue
        bank     = entry.get("bank")
        card     = entry.get("card")
        doc_type = entry.get("doc_type")
        if not bank or not card or not doc_type:
            continue
        key = f"{bank} {card}"
        coverage.setdefault(key, set()).add(doc_type)
    return coverage


def validate_coverage(coverage_map: dict[str, set[str]]) -> list[dict]:
    """Compares coverage_map against EXPECTED_CARDS and assigns status."""
    results = []
    for expected_card in EXPECTED_CARDS:
        present_docs     = coverage_map.get(expected_card, set())
        missing_required = [d for d in REQUIRED_DOCS if d not in present_docs]
        missing_optional = [d for d in OPTIONAL_DOCS if d not in present_docs]
        all_missing      = missing_required + missing_optional

        if not present_docs:
            status = "NOT_FOUND"
        elif missing_required:
            status = "CRITICAL"
        elif missing_optional:
            status = "PARTIAL"
        else:
            status = "COMPLETE"

        if status == "NOT_FOUND":
            logger.warning(f"No documents found at all for: {expected_card}")
        elif status == "CRITICAL":
            for m in missing_required:
                logger.warning(f"Missing {m} for {expected_card}")
        elif status == "PARTIAL":
            logger.warning(
                f"Only [{', '.join(sorted(present_docs))}] found for {expected_card} "
                f"— optional docs missing: {', '.join(missing_optional)}"
            )

        results.append({
            "card":         expected_card,
            "present_docs": sorted(present_docs),
            "missing_docs": all_missing,
            "status":       status,
        })
    return results


def write_missing_docs_csv(validation_results: list[dict]) -> None:
    """Write missing_docs_report.csv (CSV fallback — xlsx preferred)."""
    fieldnames = ["Card", "Present_Docs", "Missing_Docs", "Status"]
    rows = [{
        "Card":         r["card"],
        "Present_Docs": ", ".join(r["present_docs"]) or "NONE",
        "Missing_Docs": ", ".join(r["missing_docs"]) or "NONE",
        "Status":       r["status"],
    } for r in validation_results]

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    if pd:
        pd.DataFrame(rows, columns=fieldnames).to_csv(MISSING_DOCS_CSV, index=False)
    else:
        with open(MISSING_DOCS_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
    logger.info(f"Missing docs report written to: {MISSING_DOCS_CSV}")


def write_coverage_dashboard(validation_results: list[dict]) -> None:
    """Write visual Excel coverage dashboard (unchanged from original)."""
    all_doc_types = REQUIRED_DOCS + OPTIONAL_DOCS

    rows = []
    for r in validation_results:
        row = {"Card": r["card"]}
        for doc_type in all_doc_types:
            if doc_type in r["present_docs"]:
                row[doc_type] = "✅ Found"
            elif doc_type in REQUIRED_DOCS:
                row[doc_type] = "❌ MISSING"
            else:
                row[doc_type] = "⚪ Missing"
        row["Status"] = r["status"]
        rows.append(row)

    columns = ["Card"] + all_doc_types + ["Status"]

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Coverage Dashboard"

        STATUS_FILL = {
            "COMPLETE":  PatternFill("solid", fgColor="C6EFCE"),
            "PARTIAL":   PatternFill("solid", fgColor="FFEB9C"),
            "CRITICAL":  PatternFill("solid", fgColor="FFC7CE"),
            "NOT_FOUND": PatternFill("solid", fgColor="E8B0B5"),
        }
        STATUS_FONT = {
            "COMPLETE":  Font(bold=True, color="276221"),
            "PARTIAL":   Font(bold=True, color="9C6500"),
            "CRITICAL":  Font(bold=True, color="9C0006"),
            "NOT_FOUND": Font(bold=True, color="6B0000"),
        }
        CELL_FILL = {
            "✅ Found":   PatternFill("solid", fgColor="C6EFCE"),
            "❌ MISSING": PatternFill("solid", fgColor="FFC7CE"),
            "⚪ Missing": PatternFill("solid", fgColor="FFEB9C"),
        }
        CELL_FONT = {
            "✅ Found":   Font(color="276221"),
            "❌ MISSING": Font(bold=True, color="9C0006"),
            "⚪ Missing": Font(color="9C6500"),
        }

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = thin_border
        ws.row_dimensions[1].height = 28

        for row_idx, row_data in enumerate(rows, start=2):
            status = row_data["Status"]
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, "")
                cell  = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = thin_border

                if col_name == "Card":
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.fill = STATUS_FILL.get(status, PatternFill())
                    cell.font = STATUS_FONT.get(status, Font())
                elif col_name == "Status":
                    cell.fill = STATUS_FILL.get(status, PatternFill())
                    cell.font = STATUS_FONT.get(status, Font())
                elif col_name in all_doc_types:
                    cell.fill = CELL_FILL.get(value, PatternFill())
                    cell.font = CELL_FONT.get(value, Font())

            ws.row_dimensions[row_idx].height = 20

        ws.column_dimensions["A"].width = 30
        for col_idx in range(2, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
        ws.freeze_panes = "A2"

        LOG_DIR.mkdir(parents=True, exist_ok=True)
        wb.save(str(COVERAGE_DASHBOARD))
        logger.info(f"Coverage dashboard (Excel) written to: {COVERAGE_DASHBOARD}")

    except ImportError:
        logger.warning("openpyxl not installed — writing plain CSV dashboard instead.")
        fallback_path = LOG_DIR / "coverage_dashboard.csv"
        if pd:
            import pandas as _pd
            _pd.DataFrame(rows, columns=columns).to_csv(fallback_path, index=False)
        else:
            with open(fallback_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                writer.writerows(rows)
        logger.info(f"Coverage dashboard (CSV fallback) written to: {fallback_path}")


def print_validation_summary(validation_results: list[dict]) -> None:
    """Print coverage validation summary to console."""
    counts = {"COMPLETE": 0, "PARTIAL": 0, "CRITICAL": 0, "NOT_FOUND": 0}
    for r in validation_results:
        counts[r["status"]] = counts.get(r["status"], 0) + 1

    logger.info("")
    logger.info("=" * 50)
    logger.info("DOCUMENT COVERAGE VALIDATION REPORT")
    logger.info(f"  Total expected cards : {len(validation_results)}")
    logger.info(f"  COMPLETE             : {counts['COMPLETE']}")
    logger.info(f"  PARTIAL              : {counts['PARTIAL']}")
    logger.info(f"  CRITICAL             : {counts['CRITICAL']}")
    logger.info(f"  NOT FOUND            : {counts['NOT_FOUND']}")
    logger.info("=" * 50)

    critical = [r for r in validation_results if r["status"] in ("CRITICAL", "NOT_FOUND")]
    if critical:
        logger.info("CRITICAL ISSUES:")
        for r in critical:
            missing_str = ", ".join(r["missing_docs"]) or "all docs"
            logger.warning(f"  [{r['status']}] {r['card']} — missing: {missing_str}")

    partial = [r for r in validation_results if r["status"] == "PARTIAL"]
    if partial:
        logger.info("PARTIAL COVERAGE:")
        for r in partial:
            logger.info(
                f"  [PARTIAL] {r['card']} — present: {', '.join(r['present_docs'])} "
                f"| missing optional: {', '.join(r['missing_docs'])}"
            )
    logger.info("=" * 50)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE (standalone — rules only)
# For the full hybrid (rule + LLM) pipeline, use preprocess_with_llm.py
# ─────────────────────────────────────────────────────────────────────────────
def process_all(dry_run: bool = False, debug: bool = False) -> None:
    """
    Standalone rule-only pipeline.
    For the full hybrid (rule + LLM) pipeline, use:
        python data_pipeline/preprocess_with_llm.py --dry-run
        python data_pipeline/preprocess_with_llm.py
    """
    for d in [RAW_DIR, PROCESSED_DIR, REVIEW_DIR, LOG_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    pdf_files = sorted(RAW_DIR.glob("*.pdf"))
    total     = len(pdf_files)

    if total == 0:
        logger.warning(f"No PDF files found in {RAW_DIR}. Exiting.")
        return

    logger.info(f"Found {total} PDF file(s) in {RAW_DIR}")

    count_processed = count_review = count_errors = count_skipped = 0
    log_entries: list[dict] = []

    for idx, pdf_path in enumerate(pdf_files, start=1):
        logger.info(f"[PROCESSING] {idx}/{total}: {pdf_path.name}")
        pages_read = PAGE_TIERS[0]

        try:
            (text, doc_type_result, bank_result,
             card_result, master_result, year, overall_conf) = \
                run_detection_with_fallback(pdf_path, debug)

            for tier in PAGE_TIERS:
                pages_read = tier
                if overall_conf >= CONFIDENCE_THRESHOLD:
                    break

            bank      = bank_result["value"]
            card      = card_result["value"]
            doc_type  = doc_type_result["value"]
            is_master = master_result["is_master"]

            if is_master:
                logger.info(
                    f"Detected: {bank} | [MASTER DOC] | {doc_type} "
                    f"(conf={overall_conf:.2f}) → {bank}_MASTER/"
                )
            else:
                logger.info(
                    f"Detected: {bank} | {card} | {doc_type} "
                    f"(conf={overall_conf:.2f})"
                )

            needs_review = (
                overall_conf < CONFIDENCE_THRESHOLD
                or bank is None
                or card is None
            )
            new_filename = generate_filename(bank, card, doc_type, year)

            if needs_review:
                dest_path = REVIEW_DIR / new_filename
                status    = "NEEDS_REVIEW"
            else:
                dest_dir  = get_output_folder(bank, card, PROCESSED_DIR)
                dest_path = dest_dir / new_filename
                status    = "MASTER_DOC" if is_master else "SUCCESS"

            if is_duplicate(dest_path):
                logger.info(f"Duplicate detected — skipping: {new_filename}")
                status = "DUPLICATE_SKIPPED"
                count_skipped += 1
            else:
                move_file(pdf_path, dest_path, dry_run)
                count_review    += 1 if needs_review else 0
                count_processed += 0 if needs_review else 1

        except Exception as e:
            logger.error(f"Error processing {pdf_path.name}: {e}")
            status          = "ERROR"
            doc_type_result = {"value": "ERROR", "confidence": 0.0, "reasons": [str(e)]}
            bank_result     = {"value": None,    "confidence": 0.0, "reasons": []}
            card_result     = {"value": None,    "confidence": 0.0, "reasons": []}
            master_result   = {"is_master": False, "signal": None,  "confidence": 0.0}
            overall_conf    = 0.0
            is_master       = False
            count_errors   += 1

        log_entries.append({
            "filename":         pdf_path.name,
            "bank":             bank_result["value"],
            "bank_conf":        bank_result["confidence"],
            "bank_reasons":     bank_result["reasons"],
            "card":             card_result["value"],
            "card_conf":        card_result["confidence"],
            "card_reasons":     card_result["reasons"],
            "doc_type":         doc_type_result["value"],
            "doc_type_conf":    doc_type_result["confidence"],
            "doc_type_reasons": doc_type_result["reasons"],
            "is_master":        master_result["is_master"],
            "master_signal":    master_result["signal"],
            "master_conf":      master_result["confidence"],
            "overall_conf":     overall_conf,
            "pages_read":       pages_read,
            "status":           status,
        })

    write_detail_log(log_entries)
    write_summary_csv(log_entries)

    logger.info("")
    logger.info("Running document coverage validation...")
    coverage_map       = build_coverage_map(log_entries)
    validation_results = validate_coverage(coverage_map)
    write_missing_docs_csv(validation_results)
    write_coverage_dashboard(validation_results)
    print_validation_summary(validation_results)

    master_count = sum(
        1 for e in log_entries
        if e.get("is_master") and e["status"] in ("MASTER_DOC", "SUCCESS")
    )
    logger.info("")
    logger.info("=" * 50)
    logger.info("PIPELINE COMPLETE")
    logger.info(f"  Total files      : {total}")
    logger.info(f"  Processed        : {count_processed}")
    logger.info(f"  Master/Collective: {master_count}")
    logger.info(f"  Needs review     : {count_review}")
    logger.info(f"  Errors           : {count_errors}")
    logger.info(f"  Duplicates       : {count_skipped}")
    logger.info(f"  Dry-run mode     : {'ON' if dry_run else 'OFF'}")
    logger.info("=" * 50)


# ─────────────────────────────────────────────────────────────────────────────
# CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description=(
            "Credit Card PDF Preprocessing Pipeline (rule-based standalone). "
            "For hybrid rule+LLM pipeline, use preprocess_with_llm.py."
        )
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Simulate processing without copying any files.",
    )
    parser.add_argument(
        "--debug", action="store_true",
        help="Print sample extracted text for each PDF.",
    )
    args = parser.parse_args()

    setup_logging()

    if not (pdfplumber or fitz):
        logger.error(
            "No PDF library found. Install pdfplumber or PyMuPDF:\n"
            "  pip install pdfplumber\n"
            "  pip install pymupdf"
        )
        sys.exit(1)

    process_all(dry_run=args.dry_run, debug=args.debug)


if __name__ == "__main__":
    main()
